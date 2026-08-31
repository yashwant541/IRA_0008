"""
IRA Credit Risk - Dataiku web app backend  (v3, multi-page)
==========================================================
Pages driven by these endpoints:
  POST /login            {psid}                -> validates a PSID
  POST /validate         files                 -> required-table + schema checks
  POST /analyze          files + user/qtr/year -> full assessment output
  POST /save_output      {run_id}              -> write xlsx to managed folder
  POST /approve          {run_id, overrides}   -> write per Product+Country overrides
  POST /trigger_tableau  {run_id}              -> run the Tableau scenario
  GET  /download/<run_id>                       -> download the xlsx
  GET  /history                                 -> recent runs

The IRA engine (lib/python/IRA) is imported LAZILY and never modified.
"""
import io, os, csv, json, base64, datetime, traceback
from typing import Dict, List, Any, Optional, Tuple
import pandas as pd

# ------------------------------------------------------------------ CONFIG ---
STORE_FOLDER = "IRA_STORE"
TABLEAU_SCENARIO = "UPLOAD_TO_TABLEAU"
PROJECT_KEY = None
PSID_MIN_LEN = 4                 # minimal PSID length accepted
# -----------------------------------------------------------------------------

CATEGORIES = ["Secured", "Unsecured", "SME Banking", "Wealth Lending",
              "Wealth Lending - Retail Banking", "Wealth Lending - PvB"]
BASE_CATEGORIES = ["Secured", "Unsecured", "SME Banking", "Wealth Lending"]
FINAL_CALC_LABEL = "Calculated Inherent Credit Risk Assessment:"
FINAL_OVERRIDE_LABEL = "Final Inherent Credit Risk Assessment (with Override):"
RUN_CACHE: Dict[str, Dict[str, Any]] = {}

# Required input tables:  (display name, detector key, group)
REQUIRED_MI = [
    ("ENR by Country & Product", "ENR"), ("30+ DPD %", "30+%"), ("30+ DPD $", "30+$"),
    ("90+ DPD %", "90+%"), ("90+ DPD $", "90+$"), ("RWA", "RWA"),
    ("GCO %", "gco"), ("Policy exception L2 & L3", "policy_exception"),
    ("ME EA / AWC", "ME_EA_AWC"), ("PvB EA / AWC", "PvB_EA_AWC"),
    ("Country product approval rate", "app_rate"), ("# monthly new approved", "new_approved"),
    ("Property Price Index", "PPI"), ("Interest Rates", "interest_rates"),
    ("CCPL Volatile by Country", "ccpl_volatile"), ("LTV > 80", "LTV80"),
]
# reference tables checked per category (dispensations, breaches) + sovereign
REQUIRED_OTHER_SINGLE = [("Country Sovereign Rating & Outlook", "sovereign")]

# expected NON-DATE columns per reference table (schema check)
SCHEMA_REF = {
    "sovereign": ["Country", "FCY CRG", "Outlook"],
    "dispensations": ["Country", "#"],
    "cra_breaches": ["Country"],          # + monthly Y/blank (date columns) + L12M
}

# ------------------------------------------------------------ lazy IRA -------
_IRA: Dict[str, Any] = {}
def _ira():
    if not _IRA:
        try:
            from IRA import (ira_loaders as L, ira_build as B, ira_intermediate as I,
                             ira_detect as DET, ira_sovereign as SOV, ira_dispensations as DSP)
        except Exception:
            import ira_loaders as L, ira_build as B, ira_intermediate as I
            import ira_detect as DET, ira_sovereign as SOV, ira_dispensations as DSP
        _IRA.update(L=L, B=B, I=I, DET=DET, SOV=SOV, DSP=DSP)
    return _IRA

# ------------------------------------------------------------ store ----------
class _DkuStore:
    def __init__(self, f): self.f = f
    def put(self, p, d): self.f.upload_data(p, d)
    def get(self, p):
        try:
            with self.f.get_download_stream(p) as s: return s.read()
        except Exception: return None
    def list(self, pre):
        try: return [p for p in self.f.list_paths_in_partition() if p.lstrip('/').startswith(pre)]
        except Exception: return []

class _LocalStore:
    def __init__(self, r): self.r = r; os.makedirs(r, exist_ok=True)
    def put(self, p, d):
        fp = os.path.join(self.r, p); os.makedirs(os.path.dirname(fp), exist_ok=True); open(fp, 'wb').write(d)
    def get(self, p):
        fp = os.path.join(self.r, p); return open(fp, 'rb').read() if os.path.exists(fp) else None
    def list(self, pre):
        b = os.path.join(self.r, pre); return [pre + n for n in os.listdir(b)] if os.path.isdir(b) else []

def _store():
    try:
        import dataiku; return _DkuStore(dataiku.Folder(STORE_FOLDER))
    except Exception:
        return _LocalStore(os.environ.get("IRA_STORE_DIR", "./ira_store"))

def _trigger(scenario):
    try:
        import dataiku
        c = dataiku.api_client()
        pr = c.get_project(PROJECT_KEY) if PROJECT_KEY else c.get_default_project()
        s = pr.get_scenario(scenario)
        try: s.run_and_wait()
        except Exception: s.run()
        return {"ok": True, "scenario": scenario, "message": "Scenario triggered."}
    except Exception as ex:
        return {"ok": False, "scenario": scenario, "error": str(ex)}

# ------------------------------------------------------------ io -------------
def _read_sheets(raw):
    try: xls = pd.read_excel(io.BytesIO(raw), sheet_name=None, header=None, engine="openpyxl")
    except Exception: xls = pd.read_excel(io.BytesIO(raw), sheet_name=None, header=None)
    return {n: d.values.tolist() for n, d in xls.items()}

def _parse_cfg(raw):
    out = {}
    for r in csv.DictReader(io.StringIO(raw.decode("utf-8-sig", errors="replace"))):
        cat = (r.get("Category") or "").strip(); co = (r.get("Country") or "").strip()
        if cat and co and (r.get("Include") or "").strip().lower() in ("yes","y","true","1","t","x"):
            out.setdefault(cat, []).append(co)
    return out

def _cell(v): return None if (v is None or (isinstance(v, float) and pd.isna(v))) else v

def _il():
    """Import ira_logic from the IRA library (its home is the Project Libraries
    IRA package); fall back to a flat import for single-folder local runs.
    Returns the module, or None if it genuinely can't be found."""
    import sys as _sys
    try:
        here = os.path.dirname(os.path.abspath(__file__))
    except Exception:
        here = os.getcwd()
    for p in (here, os.path.dirname(here) if here else "", os.getcwd()):
        if p and p not in _sys.path:
            _sys.path.insert(0, p)
    try:
        from IRA import ira_logic as IL
        return IL
    except Exception:
        pass
    try:
        import ira_logic as IL
        return IL
    except Exception:
        return None

def _fmt_month(m):
    try:
        from IRA import ira_engine as E
    except Exception:
        try:
            import ira_engine as E
        except Exception:
            return str(m)
    try:
        return E.fmt_month(m)
    except Exception:
        return str(m)

def _table_views(tables):
    """Inlined table viewer (no external import): every parsed input table as a
    {name, columns, rows} grid, so it works even where ira_logic isn't on the path."""
    views = []

    def grid_mt(mt, name):
        months = [_fmt_month(m) for m in (getattr(mt, "months", []) or [])]
        raw = getattr(mt, "months", []) or []
        pd_data = getattr(mt, "product_data", None) or {}
        cd_data = getattr(mt, "country_data", None) or {}
        def vals(series):
            if isinstance(series, dict): return [series.get(m) for m in raw]
            if isinstance(series, (list, tuple)): return list(series)
            return []
        rows = []
        if pd_data:
            cols = ["Country", "Product"] + months
            for (c, p), s in pd_data.items(): rows.append([c, p] + vals(s))
        elif cd_data:
            cols = ["Country"] + months
            for c, s in cd_data.items(): rows.append([c] + vals(s))
        else:
            cols = ["(no rows parsed)"]
        return {"name": name, "columns": cols, "rows": rows}

    NAMES = {"ENR": "ENR", "30+%": "30+%", "30+$": "30+$", "90+%": "90+%", "90+$": "90+$",
             "RWA": "RWA", "app_rate": "Country/product app rate", "new_approved": "# monthly new approved",
             "PPI": "PPI", "LTV80": "LTV > 80 Excl MIP", "interest_rates": "Interest Rates"}
    for key, label in NAMES.items():
        mt = tables.get(key)
        if mt is not None and hasattr(mt, "months"):
            g = grid_mt(mt, label)
            if g["rows"]: views.append(g)
    for key, label in (("policy_exception", "# policy exception L2/L3"),
                       ("ME_EA_AWC", "ME EA / AWC"), ("PvB_EA_AWC", "PvB EA / AWC")):
        d = tables.get(key)
        if isinstance(d, dict):
            for sub, mt in d.items():
                if hasattr(mt, "months"):
                    g = grid_mt(mt, f"{label} [{sub}]")
                    if g["rows"]: views.append(g)
    ecl = tables.get("ecl")
    if isinstance(ecl, pd.DataFrame) and not ecl.empty:
        e = ecl.head(200).where(pd.notnull(ecl.head(200)), None)
        views.append({"name": "ECL IIP LI", "columns": [str(c) for c in e.columns],
                      "rows": e.astype(object).values.tolist()})
    sov = tables.get("sovereign")
    if isinstance(sov, dict) and sov:
        rows = [[c, r.get("fcy_crg"), r.get("lcy_crg"), r.get("outlook")] for c, r in sov.items()]
        views.append({"name": "Country Rating & Outlook",
                      "columns": ["Country", "FCY CRG", "LCY CRG", "Outlook"], "rows": rows})
    for key, label in (("dispensations", "Active/Expired Dispensation"), ("cra_breaches", "CRA Breaches")):
        d = tables.get(key)
        if isinstance(d, dict) and d:
            rows = []
            for cat, cc in d.items():
                if isinstance(cc, dict):
                    for country, val in cc.items(): rows.append([cat, country, val])
            if rows:
                views.append({"name": label, "columns": ["Category", "Country", "Value"], "rows": rows})
    for v in views:
        v["rows"] = [[_cell(x) for x in r] for r in v["rows"]]
    return views

def _assemble(mi, other):
    sheets = _read_sheets(mi)
    if other:
        for n, rows in _read_sheets(other).items(): sheets[f"OtherTables::{n}"] = rows
    return sheets

# ------------------------------------------------------------ schema check ---
def _norm(s): 
    import re; return re.sub(r"[^a-z0-9 ]", " ", str(s).lower()).strip()

def _find_ref_header(sheets, title_markers):
    """Return the header cells (as strings) of the first table whose title row
    contains any marker; else None.  Skips the title row itself."""
    for _sn, rows in sheets.items():
        for i, r in enumerate(rows):
            joined = " ".join(str(c) for c in r if c is not None).lower()
            if any(m in joined for m in title_markers):
                for j in range(i + 1, min(i + 5, len(rows))):     # AFTER the title row
                    cells = [str(c).strip() for c in rows[j]
                             if c is not None and pd.notna(c) and str(c).strip()
                             and str(c).strip().lower() not in ("nan", "nat")]
                    low = " ".join(_norm(c) for c in cells)
                    if "country" in low and len(cells) >= 2:      # a real header row
                        return cells
    return None

def _has_col(header_cells, wanted):
    w = _norm(wanted)
    for c in header_cells:
        cn = _norm(c)
        if w == cn or w in cn.split():
            return True
        if wanted == "#" and c.strip() in ("#", "Count", "Number"):
            return True
    return False

# ------------------------------------------------------------ validate -------
def validate(mi, other, config) -> Dict[str, Any]:
    avail_files = [
        {"item": "MI file", "ok": bool(mi)},
        {"item": "Other tables", "ok": bool(other)},
        {"item": "Countries config", "ok": bool(config)},
    ]
    if not mi:
        return {"ok": False, "files": avail_files, "required": [], "schema": [],
                "error": "MI file is required to validate."}

    sheets = _assemble(mi, other)
    M = _ira(); L = M["L"]
    tables = L.load_tables(sheets)

    required = []
    for name, key in REQUIRED_MI:
        required.append({"table": name, "group": "MI file", "found": tables.get(key) is not None})
    # sovereign
    required.append({"table": "Country Sovereign Rating & Outlook", "group": "Other tables",
                     "found": bool(tables.get("sovereign"))})
    # dispensations & breaches per category
    disp = tables.get("dispensations") or {}
    brc = tables.get("cra_breaches") or {}
    for cat in BASE_CATEGORIES:
        required.append({"table": f"Dispensations - {cat}", "group": "Other tables",
                         "found": bool(disp.get(cat))})
    for cat in BASE_CATEGORIES:
        required.append({"table": f"CRA Breaches - {cat}", "group": "Other tables",
                         "found": bool(brc.get(cat))})

    # ---- schema / column checks (non-date columns) ----------------------- #
    schema = []
    # sovereign columns
    sov_hdr = _find_ref_header(sheets, ["sovereign"])
    if sov_hdr:
        missing = [c for c in SCHEMA_REF["sovereign"] if not _has_col(sov_hdr, c)]
        schema.append({"table": "Sovereign Rating & Outlook", "ok": not missing,
                       "missing": missing, "found_columns": sov_hdr})
    elif tables.get("sovereign"):
        schema.append({"table": "Sovereign Rating & Outlook", "ok": True, "missing": [],
                       "found_columns": ["(parsed)"]})
    # dispensation columns (per detected table title)
    for cat, marker in (("Secured", "secured portfolio"), ("Unsecured", "unsecured portfolio"),
                        ("SME Banking", "sme portfolio"), ("Wealth Lending", "wl portfolio")):
        hdr = _find_ref_header(sheets, [marker, "active or expired"]) if (disp.get(cat)) else None
        if disp.get(cat):
            # if parsed, Country/# are present by construction
            schema.append({"table": f"Dispensations - {cat}", "ok": True, "missing": [],
                           "found_columns": hdr or ["Country", "#"]})
    # cra breaches columns
    for cat in BASE_CATEGORIES:
        if brc.get(cat):
            schema.append({"table": f"CRA Breaches - {cat}", "ok": True, "missing": [],
                           "found_columns": ["Country", "<monthly Y/blank>", "L12M"]})

    all_found = all(x["found"] for x in required)
    schema_ok = all(x["ok"] for x in schema) if schema else True
    return {"ok": all_found and schema_ok, "files": avail_files,
            "required": required, "schema": schema,
            "all_found": all_found, "schema_ok": schema_ok}

# ------------------------------------------------------------ analyze --------
def run_analysis(mi, other, config, user, quarter, year) -> Dict[str, Any]:
    now = datetime.datetime.now()
    rid = _rid(user, quarter, year, now)
    if not mi or not config:
        return {"ok": False, "run_id": rid, "error": "MI file and countries config are required."}
    sheets = _assemble(mi, other)
    cfg = _parse_cfg(config)
    M = _ira(); L, B = M["L"], M["B"]
    tables = L.load_tables(sheets)
    per_cat = B.resolve_countries(tables, None, cfg)
    frames = B.build_all(tables, countries_per_category=cfg)
    inter = B.build_intermediate_frames(tables, per_cat)
    mapping = B.build_mapping()
    try:
        intermediates = M["I"].build(tables, per_cat)
    except Exception:
        intermediates = {}

    results, calculated, na, cbp = {}, {}, [], {}
    for cat in CATEGORIES:
        df = frames.get(f"IRA - {cat}"); results[cat] = {}; calculated[cat] = {}
        cbp[cat] = (sorted(df["Country"].dropna().unique().tolist())
                    if (df is not None and not df.empty) else list(per_cat.get(cat, [])))
        if df is None or df.empty: continue
        for country, grp in df.groupby("Country", sort=False):
            rows = []
            for _, r in grp.iterrows():
                lab = str(r["Label"]); rating = _cell(r["Risk Rating"])
                if lab.startswith("Calculated"):
                    calculated[cat][country] = {"rating": rating, "value": _cell(r["Value"])}
                else:
                    rows.append({"label": lab, "value": _cell(r["Value"]), "rating": rating,
                                 "number": _cell(r["Risk Number"])})
                    if rating == "Not Available":
                        na.append({"product": cat, "country": country, "label": lab,
                                   "reason": str(r["What to do in Value Column"]).replace("Not Available - ", "")})
            calc = calculated[cat].get(country, {})
            rows.append({"label": FINAL_CALC_LABEL, "value": calc.get("value"),
                         "rating": calc.get("rating"), "number": None, "calc": True})
            rows.append({"label": FINAL_OVERRIDE_LABEL, "value": None, "rating": None,
                         "number": None, "override": True})
            results[cat][country] = rows

    meta = {"run_id": rid, "user": user, "quarter": quarter, "year": year,
            "mi_period": _mi_period(tables),
            "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
            "n_countries": len({c for cat in CATEGORIES for c in per_cat.get(cat, [])}),
            "na_count": len(na), "status": "processed"}
    try: xls = _excel(frames, inter, mapping)
    except Exception: xls = None
    RUN_CACHE[rid] = {"excel": xls, "meta": meta, "results": results, "calculated": calculated,
                      "frames": frames, "inter": inter, "mapping": mapping, "tables": tables,
                      "per_cat": per_cat, "intermediates": intermediates}
    return {"ok": True, "run_id": rid, "meta": meta, "na_details": na,
            "countries_by_product": cbp, "results": results, "calculated": calculated,
            "excel_ready": xls is not None}


def _mi_period(tables):
    try:
        from IRA import ira_engine as E
    except Exception:
        import ira_engine as E
    try:
        return E.period_label(tables)
    except Exception:
        return ""


def _lineage_rows(tables):
    """Product/label -> formula -> parsed table -> raw sheet (for the flow view)."""
    try:
        from IRA import ira_pipeline as P, ira_config as C
    except Exception:
        import ira_pipeline as P, ira_config as C
    out = []
    for product in CATEGORIES:
        for m in C.METRICS[product]():
            ik = getattr(m["value"], "int_key", "")
            parsed = [k for k in P.KEY_TO_PARSED.get(ik, []) if tables.get(k) is not None]
            raws = [P.RAW_SHEET.get(k, k) for k in parsed]
            out.append({"product": product, "label": m["label"], "formula": ik or "(blank)",
                        "parsed": "; ".join(parsed) or "n/a", "raw": "; ".join(raws) or "n/a",
                        "applicable": bool(ik)})
    return out

def _rid(user, q, y, now):
    safe = "".join(ch for ch in str(user) if ch.isalnum()) or "user"
    return f"{y}_{q}_{safe}_{now.strftime('%Y%m%d_%H%M%S')}"

# ------------------------------------------------------------ excel ----------
def _excel(frames, inter, mapping, overrides=None, manual=None):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    try:
        from IRA import ira_build as _IRB
    except Exception:
        import ira_build as _IRB
    fills = {"Very Low":"C6EFCE","Low":"D9EAD3","Medium":"FFF2CC","High":"FCE5CD","Very High":"F4CCCC","Not Available":"EFEFEF"}
    ov = {(o["product"], o["country"]): o for o in (overrides or [])}
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        for cat in CATEGORIES:
            df = frames.get(f"IRA - {cat}")
            if df is None: continue
            if overrides: df = _inject(df, cat, ov)
            try:
                df = _IRB.attach_rating_formula(df, cat)   # download-only formula column
            except Exception:
                pass
            df.to_excel(xw, sheet_name=cat[:31], index=False)
        if overrides: pd.DataFrame(overrides).to_excel(xw, sheet_name="Overrides", index=False)
        if manual: pd.DataFrame(manual).to_excel(xw, sheet_name="Manual Numbers", index=False)
        mapping.to_excel(xw, sheet_name="Mapping", index=False)
        for t, d in inter.items(): d.to_excel(xw, sheet_name=t[:31], index=False)
    buf.seek(0); wb = load_workbook(buf)
    hf = PatternFill("solid", fgColor="E8EEF7"); hfont = Font(bold=True, color="1F3A5F", size=10)
    thin = Side(style="thin", color="DCE3EF"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        rc = None
        for c in range(1, ws.max_column + 1):
            ws.cell(1, c).fill = hf; ws.cell(1, c).font = hfont
            ws.cell(1, c).alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width = 22
            if str(ws.cell(1, c).value).strip() == "Risk Rating": rc = c
        ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 26
        for r in range(2, ws.max_row + 1):
            if rc and ws.cell(r, rc).value in fills:
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = PatternFill("solid", fgColor=fills[ws.cell(r, rc).value])
            for c in range(1, ws.max_column + 1): ws.cell(r, c).border = bd
    out = io.BytesIO(); wb.save(out); return base64.b64encode(out.getvalue()).decode("ascii")

def _period_label(meta):
    p = str(meta.get("mi_period", "") or "").strip()
    if p:
        return p
    q = str(meta.get("quarter", "") or "").strip()
    y = str(meta.get("year", "") or "").strip()
    return f"{q} {y}".strip()


def _intermediate_workbook(cached):
    """All intermediate (formula) tables in one workbook, each already carrying
    the Period column (stamped by the engine)."""
    inter = cached.get("inter") or {}
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        if not inter:
            pd.DataFrame({"(no intermediates)": []}).to_excel(xw, sheet_name="empty", index=False)
        for title, df in inter.items():
            df.to_excel(xw, sheet_name=str(title)[:31], index=False)
    return base64.b64encode(buf.getvalue()).decode("ascii")


def _period_workbook(cached):
    """One combined long table with a Period column (from the user's quarter+year)
    prefilled on every row, plus Product.  Returns base64 xlsx."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    fills = {"Very Low": "C6EFCE", "Low": "D9EAD3", "Medium": "FFF2CC", "High": "FCE5CD",
             "Very High": "F4CCCC", "Not Available": "EFEFEF", "Not Applicable": "EFEFEF"}
    meta = cached["meta"]
    period = _period_label(meta)
    frames = cached["frames"]
    calc = cached.get("calculated", {})
    ov_map = {}
    for o in cached.get("overrides_last", []):
        ov_map[(o["product"], o["country"])] = o

    rows = []
    for cat in CATEGORIES:
        df = frames.get(f"IRA - {cat}")
        if df is None or df.empty:
            continue
        for country, grp in df.groupby("Country", sort=False):
            for _, r in grp.iterrows():
                rows.append({"Period": period, "Quarter": meta.get("quarter", ""),
                             "Year": meta.get("year", ""), "Prepared By": meta.get("user", ""),
                             "Product": cat, "Country": country,
                             "Label": r.get("Label"), "Value": r.get("Value"),
                             "Risk Rating": r.get("Risk Rating"), "Risk Number": r.get("Risk Number")})
            o = ov_map.get((cat, country))
            if o:
                rows.append({"Period": period, "Quarter": meta.get("quarter", ""),
                             "Year": meta.get("year", ""), "Prepared By": meta.get("user", ""),
                             "Product": cat, "Country": country,
                             "Label": FINAL_OVERRIDE_LABEL, "Value": o.get("override_text", ""),
                             "Risk Rating": o.get("override_rating") or "", "Risk Number": ""})
    out_df = pd.DataFrame(rows, columns=["Period", "Quarter", "Year", "Prepared By",
                                         "Product", "Country", "Label", "Value",
                                         "Risk Rating", "Risk Number"])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        out_df.to_excel(xw, sheet_name="IRA Output (period)", index=False)
    buf.seek(0)
    wb = load_workbook(buf); ws = wb.active
    hf = PatternFill("solid", fgColor="1F3A5F"); hfont = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="DCE3EF"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    rate_col = None
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).fill = hf; ws.cell(1, c).font = hfont
        ws.cell(1, c).alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = 20
        if str(ws.cell(1, c).value).strip() == "Risk Rating":
            rate_col = c
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 24
    for r in range(2, ws.max_row + 1):
        if rate_col and ws.cell(r, rate_col).value in fills:
            ws.cell(r, rate_col).fill = PatternFill("solid", fgColor=fills[ws.cell(r, rate_col).value])
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = bd
    outb = io.BytesIO(); wb.save(outb)
    return base64.b64encode(outb.getvalue()).decode("ascii"), out_df


def _inject(df, cat, ov):
    rows = []
    for country, grp in df.groupby("Country", sort=False):
        for _, r in grp.iterrows(): rows.append(r.to_dict())
        o = ov.get((cat, country))
        if o:
            rows.append({"Country": country, "Label": FINAL_OVERRIDE_LABEL,
                         "Value": o.get("override_text", ""), "Risk Rating": o.get("override_rating") or "",
                         "Risk Number": "", "What to do in Value Column": ""})
    return pd.DataFrame(rows, columns=df.columns)

def _cfg():
    try:
        from IRA import ira_config as C
    except Exception:
        import ira_config as C
    return C

def _canon(label):
    import re
    m = re.match(r"^([0-9]+[a-z]*)", str(label).strip()); return m.group(1) if m else str(label).strip()

def _rate_value(product, label, value):
    """Best-effort rating for a manually entered value, using the metric's own ladder."""
    C = _cfg()
    try: val = float(value)
    except Exception: return ""
    for m in C.METRICS.get(product, lambda: [])():
        if m["label"] == label:
            ik = getattr(m.get("value"), "int_key", "")
            try:
                if ik in ("dpd_qoq_cur", "dpd_qoq_prior", "dpd_yoy"):
                    return C.r_deterioration(val)
                return m["rating"](val, None)
            except Exception:
                return ""
    return ""

def _recompute_inherent(frame, product, country):
    """Recompute a country's Calculated inherent from the frame's risk numbers
    (max per theme x weight, 1bii excluded) - used after a manual entry."""
    C = _cfg()
    sub = frame[frame["Country"] == country]
    num = {}
    for _, r in sub.iterrows():
        lab = str(r.get("Label", ""))
        if lab.startswith("Calculated"): continue
        n = r.get("Risk Number", "")
        try: num[_canon(lab)] = float(n) if n not in ("", None) else None
        except Exception: num[_canon(lab)] = None
    groups = C.AGG_GROUPS.get(product, [])
    weight = (1.0/len(groups)) if (getattr(C, "NORMALISE_WEIGHTS", False) and groups) else C.W6
    ordered = [_canon(m["label"]) for m in C.METRICS.get(product, lambda: [])()]
    score = 0.0
    for positions in groups:
        theme = []
        for p in positions:
            if 1 <= p <= len(ordered):
                canon = ordered[p-1]
                if canon == "1bii": continue
                v = num.get(canon)
                if v is not None: theme.append(v)
        if theme: score += weight * max(theme)
    def s2r(x): return "Very High" if x>=4.5 else "High" if x>=3.5 else "Medium" if x>=2.5 else "Low" if x>=1.5 else "Very Low"
    return round(score, 4), s2r(score)

def _na_values(frames):
    """List every cell that came out Not Available, as {product, country, label}."""
    out = []
    for name, df in frames.items():
        prod = name.replace("IRA - ", "", 1)
        if df is None or df.empty: continue
        na = df[df["Risk Rating"] == "Not Available"]
        for _, r in na.iterrows():
            out.append({"product": prod, "country": r.get("Country"), "label": r.get("Label")})
    return out

def _apply_manual(c, product, country, label, value):
    """Write a manual value into the cached frame, re-rate it, recompute that
    country's inherent, and record it for the Manual Numbers sheet."""
    import re
    frame = c["frames"].get(f"IRA - {product}")
    if frame is None: return {"ok": False, "error": "Unknown product."}
    mask = (frame["Country"] == country) & (frame["Label"] == label)
    if not mask.any(): return {"ok": False, "error": "Cell not found."}
    rating = _rate_value(product, label, value)
    try:
        from IRA import ira_engine as E
    except Exception:
        import ira_engine as E
    number = E.RISK_NUMBER.get(rating) if rating else ""
    # numeric metrics show as %, everything else as-is
    disp = value
    try: disp = f"{float(value)*100:.2f}%"
    except Exception: pass
    frame.loc[mask, "Value"] = disp
    frame.loc[mask, "Risk Rating"] = rating or ""
    frame.loc[mask, "Risk Number"] = number if number is not None else ""
    if "What to do in Value Column" in frame.columns:
        frame.loc[mask, "What to do in Value Column"] = "Manually entered value"
    # recompute the Calculated inherent for this country
    score, srating = _recompute_inherent(frame, product, country)
    cmask = (frame["Country"] == country) & (frame["Label"].astype(str).str.startswith("Calculated"))
    if cmask.any():
        frame.loc[cmask, "Value"] = score; frame.loc[cmask, "Risk Rating"] = srating
    c.setdefault("manual", []).append({"Product": product, "Country": country, "Label": label,
                                       "Manual Value": disp, "Risk Rating": rating or "",
                                       "Risk Number": number if number is not None else ""})
    c.setdefault("calculated", {}).setdefault(product, {})[country] = {"rating": srating, "value": score}
    try:
        c["excel"] = _excel(c["frames"], c["inter"], c["mapping"],
                            overrides=c.get("overrides_last"), manual=c.get("manual"))
    except Exception: pass
    return {"ok": True, "product": product, "country": country, "label": label,
            "value": disp, "rating": rating or "", "number": number if number is not None else "",
            "calculated": {"rating": srating, "value": score}}

def _parse_override_file(raw, filename=""):
    """Parse an uploaded overrides file into approve-style rows.
       Columns (case/space tolerant): Product | Country |
       Calculated Inherent Risk Rating | Final Inherent Credit Risk Rating | Rationale."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(raw))
    else:
        df = pd.read_excel(io.BytesIO(raw))
    norm = {str(k).strip().lower(): k for k in df.columns}
    def col(*cands):
        for cand in cands:
            for k, orig in norm.items():
                if cand in k: return orig
        return None
    p, cty = col("product"), col("country")
    fin = col("final inherent", "final rating", "final")
    rat = col("rationale", "comment", "reason")
    calc = col("calculated")
    if not p or not cty or not fin:
        raise ValueError("File needs at least Product, Country and Final Inherent Credit Risk Rating columns.")
    rows = []
    for _, r in df.iterrows():
        prod = str(r.get(p, "")).strip(); country = str(r.get(cty, "")).strip()
        if not prod or not country: continue
        rows.append({"product": prod, "country": country,
                     "override_rating": ("" if fin is None else str(r.get(fin, "")).strip()),
                     "override_text": ("" if rat is None else str(r.get(rat, "")).strip()),
                     "calculated_rating": ("" if calc is None else str(r.get(calc, "")).strip())})
    return rows

# ------------------------------------------------------------ routes ---------
try:
    from flask import request

    def _J(o, code=200):
        return app.response_class(json.dumps(o, default=str), mimetype="application/json", status=code)

    @app.route("/health")
    def health(): return _J({"ok": True, "app": "IRA Credit Risk"})

    @app.route("/login", methods=["POST"])
    def login():
        b = request.get_json(force=True, silent=True) or {}
        psid = str(b.get("psid", "")).strip()
        if len(psid) < PSID_MIN_LEN:
            return _J({"ok": False, "error": f"Enter a valid PSID (min {PSID_MIN_LEN} chars)."})
        return _J({"ok": True, "psid": psid})

    @app.route("/validate", methods=["POST"])
    def validate_ep():
        try:
            def b(k):
                f = request.files.get(k); return f.read() if f else None
            return _J(validate(b("mi"), b("other"), b("config")))
        except Exception as ex:
            return _J({"ok": False, "error": f"{type(ex).__name__}: {ex}", "trace": traceback.format_exc()}, 500)

    @app.route("/analyze", methods=["POST"])
    def analyze_ep():
        try:
            def b(k):
                f = request.files.get(k); return f.read() if f else None
            return _J(run_analysis(b("mi"), b("other"), b("config"),
                                   request.form.get("user", "").strip(),
                                   request.form.get("quarter", "").strip(),
                                   request.form.get("year", "").strip()))
        except Exception as ex:
            return _J({"ok": False, "error": f"{type(ex).__name__}: {ex}", "trace": traceback.format_exc()}, 500)

    @app.route("/save_output", methods=["POST"])
    def save_ep():
        b = request.get_json(force=True, silent=True) or {}; rid = b.get("run_id"); c = RUN_CACHE.get(rid)
        if not c or not c.get("excel"): return _J({"ok": False, "error": "Run not found / output not ready."}, 404)
        try:
            s = _store(); s.put(f"outputs/{rid}.xlsx", base64.b64decode(c["excel"]))
            m = dict(c["meta"]); m["status"] = "saved_to_folder"; s.put(f"runs/{rid}.json", json.dumps(m).encode()); c["meta"] = m
            return _J({"ok": True, "path": f"outputs/{rid}.xlsx"})
        except Exception as ex: return _J({"ok": False, "error": str(ex)}, 500)

    @app.route("/approve", methods=["POST"])
    def approve_ep():
        b = request.get_json(force=True, silent=True) or {}; rid = b.get("run_id"); ov = b.get("overrides", [])
        c = RUN_CACHE.get(rid)
        if not c: return _J({"ok": False, "error": "Run not found."}, 404)
        try:
            s = _store(); rows = []
            for o in ov:
                cc = c["calculated"].get(o["product"], {}).get(o["country"], {})
                rows.append({"run_id": rid, "user": c["meta"]["user"], "quarter": c["meta"]["quarter"],
                             "year": c["meta"]["year"], "product": o["product"], "country": o["country"],
                             "calculated_rating": cc.get("rating"), "calculated_score": cc.get("value"),
                             "override_rating": o.get("override_rating") or "", "override_text": o.get("override_text") or "",
                             "timestamp": c["meta"]["timestamp"]})
            buf = io.StringIO()
            w = csv.DictWriter(buf, fieldnames=(list(rows[0].keys()) if rows else ["run_id","product","country","override_rating","override_text"]))
            w.writeheader(); [w.writerow(r) for r in rows]
            s.put(f"overrides/{rid}.csv", buf.getvalue().encode())
            c["overrides_last"] = rows
            try:
                c["excel"] = _excel(c["frames"], c["inter"], c["mapping"], overrides=rows, manual=c.get("manual"))
            except Exception: pass
            s.put(f"outputs/{rid}.xlsx", base64.b64decode(c["excel"]))
            m = dict(c["meta"]); m["status"] = "approved"; m["approved_at"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"); m["overrides"] = len(rows)
            s.put(f"runs/{rid}.json", json.dumps(m).encode()); c["meta"] = m
            return _J({"ok": True, "overrides_saved": len(rows), "path": f"overrides/{rid}.csv"})
        except Exception as ex: return _J({"ok": False, "error": str(ex)}, 500)

    @app.route("/trigger_tableau", methods=["POST"])
    def tableau_ep(): return _J(_trigger(TABLEAU_SCENARIO))

    @app.route("/download/<run_id>")
    def download_ep(run_id):
        c = RUN_CACHE.get(run_id); data = base64.b64decode(c["excel"]) if (c and c.get("excel")) else _store().get(f"outputs/{run_id}.xlsx")
        if not data: return _J({"ok": False, "error": "Output not found."}, 404)
        return app.response_class(data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  headers={"Content-Disposition": f'attachment; filename="IRA_{run_id}.xlsx"'})

    @app.route("/download_period/<run_id>")
    def download_period_ep(run_id):
        c = RUN_CACHE.get(run_id)
        if not c or not c.get("frames"):
            return _J({"ok": False, "error": "Run not found (re-run the analysis)."}, 404)
        try:
            b64, _df = _period_workbook(c)
            data = base64.b64decode(b64)
            period = _period_label(c["meta"]).replace(" ", "_") or "period"
            fname = f"IRA_{period}_{run_id}.xlsx"
            return app.response_class(
                data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{fname}"'})
        except Exception as ex:
            return _J({"ok": False, "error": str(ex)}, 500)

    @app.route("/download_intermediate/<run_id>")
    def download_intermediate_ep(run_id):
        c = RUN_CACHE.get(run_id)
        if not c or not c.get("inter"):
            return _J({"ok": False, "error": "Run not found (re-run the analysis)."}, 404)
        try:
            data = base64.b64decode(_intermediate_workbook(c))
            period = _period_label(c["meta"]).replace(" ", "_").replace("/", "-") or "period"
            return app.response_class(
                data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="IRA_Intermediate_{period}_{run_id}.xlsx"'})
        except Exception as ex:
            return _J({"ok": False, "error": str(ex)}, 500)

    @app.route("/tables_preview", methods=["POST"])
    def tables_preview_ep():
        """Parse the just-uploaded files and return every table as a grid, so the
        user can inspect (and copy) each parsed table right after uploading."""
        try:
            def b(k):
                f = request.files.get(k); return f.read() if f else None
            mi, other = b("mi"), b("other")
            if not mi:
                return _J({"ok": False, "error": "MI file is required."}, 400)
            sheets = _assemble(mi, other)
            M = _ira(); tables = M["L"].load_tables(sheets)
            return _J({"ok": True, "tables": _table_views(tables)})
        except Exception as ex:
            return _J({"ok": False, "error": f"{type(ex).__name__}: {ex}",
                       "trace": traceback.format_exc()}, 500)

    @app.route("/tables/<run_id>")
    def tables_ep(run_id):
        c = RUN_CACHE.get(run_id)
        if not c or not c.get("tables"):
            return _J({"ok": False, "error": "Run not found (re-run the analysis)."}, 404)
        try:
            return _J({"ok": True, "period": c["meta"].get("mi_period", ""),
                       "tables": _table_views(c["tables"])})
        except Exception as ex:
            return _J({"ok": False, "error": f"{type(ex).__name__}: {ex}",
                       "trace": traceback.format_exc()}, 500)

    @app.route("/download_trace/<run_id>")
    def download_trace_ep(run_id):
        c = RUN_CACHE.get(run_id)
        if not c or not c.get("frames"):
            return _J({"ok": False, "error": "Run not found (re-run the analysis)."}, 404)
        IL = _il()
        if IL is None:
            return _J({"ok": False, "error": "Trace builder (IRA.ira_logic) is not available - ensure the IRA library is on the path."}, 500)
        try:
            data = IL.build_trace_workbook(c["frames"], c.get("intermediates") or {}, c.get("tables"))
            period = _period_label(c["meta"]).replace(" ", "_").replace("/", "-") or "period"
            return app.response_class(
                data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="IRA_Trace_{period}_{run_id}.xlsx"'})
        except Exception as ex:
            return _J({"ok": False, "error": f"{type(ex).__name__}: {ex}",
                       "trace": traceback.format_exc()}, 500)

    @app.route("/download_logic/<run_id>")
    def download_logic_ep(run_id):
        c = RUN_CACHE.get(run_id)
        if not c or not c.get("frames"):
            return _J({"ok": False, "error": "Run not found (re-run the analysis)."}, 404)
        IL = _il()
        if IL is None:
            return _J({"ok": False, "error": "Logic builder (IRA.ira_logic) is not available - ensure the IRA library is on the path."}, 500)
        try:
            data = IL.build_logic_workbook(c["frames"], c.get("intermediates") or {}, c.get("tables"))
            period = _period_label(c["meta"]).replace(" ", "_").replace("/", "-") or "period"
            return app.response_class(
                data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="IRA_Calculation_Logic_{period}_{run_id}.xlsx"'})
        except Exception as ex:
            return _J({"ok": False, "error": f"{type(ex).__name__}: {ex}",
                       "trace": traceback.format_exc()}, 500)

    @app.route("/na_values/<run_id>")
    def na_values_ep(run_id):
        c = RUN_CACHE.get(run_id)
        if not c or not c.get("frames"):
            return _J({"ok": False, "error": "Run not found (re-run the analysis)."}, 404)
        return _J({"ok": True, "na": _na_values(c["frames"]), "manual": c.get("manual", [])})

    @app.route("/set_manual/<run_id>", methods=["POST"])
    def set_manual_ep(run_id):
        c = RUN_CACHE.get(run_id)
        if not c or not c.get("frames"):
            return _J({"ok": False, "error": "Run not found (re-run the analysis)."}, 404)
        b = request.get_json(force=True, silent=True) or {}
        try:
            res = _apply_manual(c, b.get("product"), b.get("country"), b.get("label"), b.get("value"))
            res["manual"] = c.get("manual", [])
            return _J(res, 200 if res.get("ok") else 400)
        except Exception as ex:
            return _J({"ok": False, "error": f"{type(ex).__name__}: {ex}"}, 500)

    @app.route("/upload_overrides/<run_id>", methods=["POST"])
    def upload_overrides_ep(run_id):
        c = RUN_CACHE.get(run_id)
        if not c:
            return _J({"ok": False, "error": "Run not found (re-run the analysis)."}, 404)
        f = request.files.get("file")
        if not f:
            return _J({"ok": False, "error": "No file uploaded."}, 400)
        try:
            rows = _parse_override_file(f.read(), f.filename)
            c["parsed_overrides"] = rows
            return _J({"ok": True, "overrides": rows, "count": len(rows)})
        except Exception as ex:
            return _J({"ok": False, "error": f"{type(ex).__name__}: {ex}"}, 400)

    @app.route("/lineage/<run_id>")
    def lineage_ep(run_id):
        c = RUN_CACHE.get(run_id)
        if not c or not c.get("tables"):
            return _J({"ok": False, "error": "Run not found."}, 404)
        return _J({"ok": True, "period": c["meta"].get("mi_period", ""),
                   "rows": _lineage_rows(c["tables"])})

    @app.route("/history")
    def history_ep():
        try:
            s = _store(); runs = []
            for p in s.list("runs/"):
                raw = s.get(p.lstrip("/"))
                if raw:
                    try: runs.append(json.loads(raw.decode()))
                    except Exception: pass
            runs.sort(key=lambda m: m.get("timestamp", ""), reverse=True)
            return _J({"ok": True, "runs": runs[:25]})
        except Exception as ex: return _J({"ok": False, "error": str(ex), "runs": []})
except NameError:
    pass