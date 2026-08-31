"""
ira_detect.py
=============
A standalone, defensive "what did we detect?" pass over an input workbook.

For every sheet it reports:
  * whether a table was DETECTED
  * the detected SHAPE (country+product block, side-by-side L2|L3, 2x2 quadrant,
    stacked country-only, plain country-only, horizontal codes, fx, long)
  * how many months / countries / products were found, and the month range
  * a simplified, cleaned version of the table(s) with Mar-YY headers and
    normalised product names

`write_detection_report(sheets, path)` writes it all to one Excel:
  - a "Detection Summary" sheet (one row per detected table)
  - one sheet per detected table holding the cleaned data

Nothing here raises: a sheet that can't be parsed is reported as NOT DETECTED
with the reason, so the report always completes.
"""

from __future__ import annotations
from typing import Dict, List, Any, Tuple
import pandas as pd

try:
    from . import ira_engine as E
    from . import ira_registry as R
    from . import ira_normalize as N
except ImportError:
    import ira_engine as E
    import ira_registry as R
    import ira_normalize as N


# --------------------------------------------------------------------------- #
#  classification helpers
# --------------------------------------------------------------------------- #
def _date_rows(rows) -> List[int]:
    return [i for i, r in enumerate(rows) if len(E._month_cols(r)) >= 2]


def _has_product_subrows(rows) -> bool:
    for r in rows:
        c = r[0] if r else None
        if E.is_product_label(c):
            return True
    return False


def _vertical_gap_col(rows) -> bool:
    """A fully-blank column with real columns on both sides (side-by-side)."""
    width = max((len(r) for r in rows), default=0)
    grid = [list(r) + [None] * (width - len(r)) for r in rows]
    for j in range(1, width - 1):
        if all(E._blank(row[j]) for row in grid):
            left = any(not E._blank(row[k]) for row in grid for k in range(j))
            right = any(not E._blank(row[k]) for row in grid for k in range(j + 1, width))
            if left and right:
                return True
    return False


def _count_titles(rows) -> int:
    """Number of section-title rows (non-date, non-'Country' text with a blank
    data area) - indicates stacked sub-tables."""
    n = 0
    for r in rows:
        c = r[0] if r else None
        if isinstance(c, str) and c.strip() and "country" not in c.lower() \
                and not E._row_has_data(r) and len(E._month_cols(r)) == 0:
            n += 1
    return n


def classify(rows) -> str:
    """Best-guess shape for a sheet, independent of its name."""
    drows = _date_rows(rows)
    if not drows:
        # no month columns - fx / ccpl / long
        flat = [str(c).strip().lower() for r in rows for c in r if c is not None]
        if any(x in ("global", "kr", "hk") for x in flat) and "volatile" in flat:
            return R.SHAPE_CCPL_HORIZONTAL
        if any("rate" in x or "currency" in x or "code" in x for x in flat):
            return R.SHAPE_FX
        return R.SHAPE_LONG
    if _vertical_gap_col(rows) and _has_product_subrows(rows):
        return R.SHAPE_SIDE_BY_SIDE
    if _vertical_gap_col(rows) and not _has_product_subrows(rows):
        return R.SHAPE_QUADRANT
    if _has_product_subrows(rows):
        return R.SHAPE_PRODUCT_BLOCK
    if _count_titles(rows) >= 2:
        return R.SHAPE_STACKED
    return R.SHAPE_COUNTRY_ONLY


# --------------------------------------------------------------------------- #
#  build simplified tables + facts per shape
# --------------------------------------------------------------------------- #
def _country_only_df(tbl: "E.MonthTable") -> pd.DataFrame:
    hdrs = E.fmt_months(tbl.months)
    rows = []
    for c, s in tbl.country_data.items():
        row = {"Row": c}
        for m, h in zip(tbl.months, hdrs):
            row[h] = s.get(m)
        rows.append(row)
    return pd.DataFrame(rows)


def detect_sheet(name: str, rows: List[List[Any]]) -> List[dict]:
    """Return one or more detection records for a sheet."""
    # is it effectively empty?
    if not rows or all(E._blank(c) for r in rows for c in r):
        return [_rec(name, "", "(empty)", detected=False,
                     notes="Sheet is present but contains no data.")]

    # prefer the registry's expected shape if the sheet name matches
    exp = _match_registry(name)
    shape = exp.shape if exp else classify(rows)

    try:
        if shape == R.SHAPE_PRODUCT_BLOCK:
            tbl = E.parse_country_product_block(rows)
            df = N.block_to_long(tbl)
            prods = sorted({p for (_c, p) in tbl.product_data})
            return [_rec(name, shape, name, detected=bool(tbl.months and df.shape[0]),
                         months=tbl.months,
                         countries=[c for c in tbl.countries()
                                    if not str(c).lower().startswith("total")],
                         products=[E.PRODUCT_OUT_NAME.get(p, p) for p in prods],
                         df=df,
                         notes=_note_block(tbl, prods))]

        if shape == R.SHAPE_SIDE_BY_SIDE:
            d = E.parse_side_by_side(rows)
            out = []
            for side, label in (("left", "L2"), ("right", "L3")):
                t = d.get(side)
                if t is None:
                    out.append(_rec(name, shape, f"{name} [{label}]",
                                    detected=False,
                                    notes=f"{label} block not found (split column "
                                          "not detected)."))
                    continue
                df = N.block_to_long(t)
                prods = sorted({p for (_c, p) in t.product_data})
                out.append(_rec(name, shape, f"{name} [{label}]",
                                detected=bool(df.shape[0]), months=t.months,
                                countries=t.countries(),
                                products=[E.PRODUCT_OUT_NAME.get(p, p) for p in prods],
                                df=df,
                                notes=f"Side {label} of the L2|L3 pair."))
            return out

        if shape == R.SHAPE_QUADRANT:
            q = E.parse_gco_quadrants(rows)
            out = []
            for prod, t in q.items():
                df = _country_only_df(t)
                out.append(_rec(name, shape, f"{name} [{E.PRODUCT_OUT_NAME.get(prod, prod)}]",
                                detected=bool(df.shape[0]), months=t.months,
                                countries=list(t.country_data.keys()),
                                products=[E.PRODUCT_OUT_NAME.get(prod, prod)],
                                df=df,
                                notes="One quadrant of the GCO 2x2 grid."))
            if not out:
                out = [_rec(name, shape, name, detected=False,
                            notes="No quadrants detected.")]
            return out

        if shape == R.SHAPE_STACKED:
            st = E.parse_stacked_country_only(rows)
            out = []
            for title, t in st.items():
                df = _country_only_df(t)
                out.append(_rec(name, shape, f"{name} [{title}]",
                                detected=bool(df.shape[0]), months=t.months,
                                countries=list(t.country_data.keys()),
                                df=df, notes=f"Stacked sub-table: {title}"))
            if not out:
                out = [_rec(name, shape, name, detected=False,
                            notes="No stacked sub-tables detected.")]
            return out

        if shape == R.SHAPE_COUNTRY_ONLY:
            t = E.parse_country_only(rows)
            df = _country_only_df(t)
            return [_rec(name, shape, name, detected=bool(df.shape[0]),
                         months=t.months, countries=list(t.country_data.keys()),
                         df=df, notes="Country/row-keyed monthly table.")]

        if shape == R.SHAPE_CCPL_HORIZONTAL:
            d = E.parse_ccpl_volatile(rows)
            df = pd.DataFrame([{"Code": k, "Value": v} for k, v in d.items()])
            return [_rec(name, shape, name, detected=bool(len(d)),
                         countries=list(d.keys()), df=df,
                         notes="Horizontal country-code table.")]

        if shape == R.SHAPE_FX:
            d = E.parse_fx(rows)
            df = pd.DataFrame([{"Currency": k, "Rate": v} for k, v in d.items()])
            return [_rec(name, shape, name, detected=bool(len(d)), df=df,
                         notes="Currency -> rate reference.")]

        # long / unknown
        df = E.parse_long(rows)
        return [_rec(name, R.SHAPE_LONG, name, detected=not df.empty,
                     df=df.head(50),
                     notes="Tidy/long table (first 50 rows shown)."
                           if not df.empty else "Could not parse as a table.")]

    except Exception as ex:
        return [_rec(name, shape, name, detected=False,
                     notes=f"Parser error: {type(ex).__name__}: {ex}")]


def _note_block(tbl, prods) -> str:
    msgs = []
    if "Other" in prods:
        msgs.append("'Other' rows present (ignored).")
    if tbl.grand_total:
        msgs.append("Has a Total row.")
    miss = [E.PRODUCT_OUT_NAME[c] for c in
            ("Consumer Secured", "Consumer Unsecured", "SME Banking", "Wealth Banking")
            if c not in prods]
    if miss:
        msgs.append("Missing products: " + ", ".join(miss))
    return " ".join(msgs) or "All four products present."


def _match_registry(name: str):
    norm = lambda s: "".join(ch for ch in str(s).lower() if ch.isalnum() or ch in "%$")
    for e in R.REGISTRY:
        if e.fabricated:
            continue
        if any(norm(a) == norm(name) for a in e.aliases):
            return e
    return None


def _rec(sheet, shape, title, detected, months=None, countries=None,
         products=None, df=None, notes=""):
    months = months or []
    countries = countries or []
    products = products or []
    return {
        "sheet": sheet, "shape": shape, "title": title, "detected": detected,
        "n_months": len(months),
        "month_range": (f"{E.fmt_month(months[0])} - {E.fmt_month(months[-1])}"
                        if months else ""),
        "n_countries": len([c for c in countries
                            if not str(c).lower().startswith("total")]),
        "countries": ", ".join(str(c) for c in countries[:10]),
        "n_products": len(products),
        "products": ", ".join(str(p) for p in products),
        "notes": notes,
        "df": df if df is not None else pd.DataFrame(),
    }


def detect_all(sheets: Dict[str, List[List[Any]]]) -> List[dict]:
    out = []
    for name, rows in sheets.items():
        out.extend(detect_sheet(name, rows))
    return out


# --------------------------------------------------------------------------- #
#  Excel report
# --------------------------------------------------------------------------- #
def write_detection_report(sheets: Dict[str, List[List[Any]]], path: str) -> List[dict]:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    records = detect_all(sheets)

    summary = pd.DataFrame([{
        "Sheet": r["sheet"],
        "Detected": "Yes" if r["detected"] else "NO",
        "Detected shape": r["shape"],
        "Table": r["title"],
        "# Months": r["n_months"],
        "Month range": r["month_range"],
        "# Countries": r["n_countries"],
        "Countries": r["countries"],
        "# Products": r["n_products"],
        "Products": r["products"],
        "Notes": r["notes"],
    } for r in records])

    used = set()
    name_map = []
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        summary.to_excel(xw, sheet_name="Detection Summary", index=False)
        used.add("Detection Summary")
        for r in records:
            if r["df"] is None or r["df"].empty:
                name_map.append(None)
                continue
            sn = E.sanitize_sheet_name(r["title"], used)
            name_map.append(sn)
            r["df"].to_excel(xw, sheet_name=sn, index=False)

    # formatting
    wb = load_workbook(path)
    hfill = PatternFill("solid", fgColor="1F4E78")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    ok = PatternFill("solid", fgColor="C6EFCE")
    bad = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb["Detection Summary"]
    widths = [26, 10, 22, 30, 9, 20, 11, 34, 10, 30, 46]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).fill = hfill
        ws.cell(1, c).font = hfont
        ws.cell(1, c).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    for row in range(2, ws.max_row + 1):
        det = ws.cell(row, 2).value
        for c in range(1, ws.max_column + 1):
            ws.cell(row, c).border = border
            ws.cell(row, c).font = Font(name="Arial", size=9)
            ws.cell(row, c).alignment = Alignment(vertical="top",
                                                  wrap_text=(c in (8, 10, 11)))
        ws.cell(row, 2).fill = ok if det == "Yes" else bad
        ws.cell(row, 2).font = Font(name="Arial", size=9, bold=True,
                                    color="006100" if det == "Yes" else "9C0006")
    ws.freeze_panes = "A2"

    for ws2 in wb.worksheets:
        if ws2.title == "Detection Summary":
            continue
        for c in range(1, ws2.max_column + 1):
            ws2.cell(1, c).fill = hfill
            ws2.cell(1, c).font = hfont
            ws2.column_dimensions[get_column_letter(c)].width = 16
        ws2.freeze_panes = "A2"
    wb.save(path)
    return records
