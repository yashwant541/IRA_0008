"""
ira_diagnostics.py
==================
Audits an input workbook against the expected-table REGISTRY and produces a
"meta output" describing, for every table, exactly what is present, missing, or
mis-read - and why.

Public API
----------
    records, summary = audit(sheets)         # sheets = {sheet_name: rows}
    write_report(records, summary, path)     # -> formatted .xlsx
    print_summary(records, summary)          # console

Each record is a dict with:
    key, display, status, severity, matched_sheet, shape, months, n_months,
    month_range, n_countries, countries, n_products, products, used_by, issues

Status values:  OK · EMPTY · MISSING · COLLISION · SHAPE_ISSUE · FABRICATED
Severity:       OK · INFO · WARN · ERROR   (drives colour + sort order)
"""

from __future__ import annotations
from typing import Dict, List, Any, Tuple
from collections import defaultdict

try:                        # inside the IRA package (Dataiku library)
    from . import ira_engine as E
    from . import ira_registry as R
except ImportError:         # flat / standalone
    import ira_engine as E
    import ira_registry as R


SEV_ORDER = {"ERROR": 0, "WARN": 1, "INFO": 2, "OK": 3}


# --------------------------------------------------------------------------- #
#  Sheet resolution (mirrors the loader's tolerant matching, with collision
#  detection so we can SEE when two expected tables grab the same sheet).
# --------------------------------------------------------------------------- #
def _norm(s: str) -> str:
    return "".join(ch for ch in str(s).lower() if ch.isalnum() or ch in "%$")


def _resolve(sheets: Dict[str, List[List[Any]]]) -> Tuple[dict, dict]:
    """
    Returns (resolved, sheet_usage):
        resolved     : expected_key -> matched sheet name (or None)
        sheet_usage  : sheet name   -> list of expected keys that matched it
    """
    norm_index = defaultdict(list)
    for name in sheets:
        norm_index[_norm(name)].append(name)

    resolved: Dict[str, Any] = {}
    sheet_usage: Dict[str, List[str]] = defaultdict(list)
    for e in R.REGISTRY:
        if e.fabricated:
            resolved[e.key] = None
            continue
        match = None
        for alias in e.aliases:
            cands = norm_index.get(_norm(alias))
            if cands:
                match = cands[0]
                break
        resolved[e.key] = match
        if match:
            sheet_usage[match].append(e.key)
    return resolved, sheet_usage


# --------------------------------------------------------------------------- #
#  Per-shape parse + validate
# --------------------------------------------------------------------------- #
def _validate(e: "R.Expected", rows: List[List[Any]]) -> dict:
    """Parse `rows` per the expected shape and return facts + issues."""
    facts = dict(shape=e.shape, months=[], n_months=0, month_range="",
                 n_countries=0, countries=[], n_products=0, products=[],
                 issues=[], parsed_ok=True)

    def _rangestr(months):
        return f"{months[0]} - {months[-1]}" if months else ""

    try:
        if e.shape == R.SHAPE_PRODUCT_BLOCK:
            t = E.parse_country_product_block(rows)
            countries = [c for c in t.countries()
                         if not str(c).lower().startswith("total")]
            products = sorted({p for (_c, p) in t.product_data})
            facts.update(months=t.months, n_months=len(t.months),
                         month_range=_rangestr(t.months),
                         countries=countries, n_countries=len(countries),
                         products=products, n_products=len(products))
            if e.expects_products and not products:
                facts["issues"].append("No product sub-rows detected under countries.")

        elif e.shape == R.SHAPE_COUNTRY_ONLY:
            t = E.parse_country_only(rows)
            cs = list(t.country_data.keys())
            facts.update(months=t.months, n_months=len(t.months),
                         month_range=_rangestr(t.months),
                         countries=cs, n_countries=len(cs))

        elif e.shape == R.SHAPE_SIDE_BY_SIDE:
            d = E.parse_side_by_side(rows)
            if "right" not in d:
                facts["issues"].append("Only ONE block found - the L2|L3 split "
                                       "column was not detected.")
            left = d.get("left")
            months = left.months if left else []
            products = sorted({p for (_c, p) in left.product_data}) if left else []
            facts.update(months=months, n_months=len(months),
                         month_range=_rangestr(months),
                         countries=(left.countries() if left else []),
                         n_countries=(len(left.countries()) if left else 0),
                         products=products, n_products=len(products))
            facts["blocks"] = [k for k in ("left", "right") if k in d]

        elif e.shape == R.SHAPE_QUADRANT:
            q = E.parse_gco_quadrants(rows)
            prods = list(q.keys())
            facts.update(products=prods, n_products=len(prods))
            example = next(iter(q.values()), None)
            if example:
                facts.update(months=example.months, n_months=len(example.months),
                             month_range=_rangestr(example.months),
                             countries=list(example.country_data.keys()),
                             n_countries=len(example.country_data))
            if len(prods) < 4:
                facts["issues"].append(
                    f"Expected 4 product quadrants, detected {len(prods)}: {prods}")

        elif e.shape == R.SHAPE_STACKED:
            st = E.parse_stacked_country_only(rows)
            facts["sections"] = list(st.keys())
            example = next(iter(st.values()), None)
            if example:
                facts.update(months=example.months, n_months=len(example.months),
                             month_range=_rangestr(example.months),
                             countries=list(example.country_data.keys()),
                             n_countries=len(example.country_data))
            facts["n_products"] = len(st)
            facts["products"] = list(st.keys())
            if not st:
                facts["issues"].append("No stacked sub-tables detected.")

        elif e.shape == R.SHAPE_CCPL_HORIZONTAL:
            d = E.parse_ccpl_volatile(rows)
            facts.update(countries=list(d.keys()), n_countries=len(d))
            if not d:
                facts["issues"].append("No country codes / volatile row detected.")

        elif e.shape == R.SHAPE_FX:
            d = E.parse_fx(rows)
            facts.update(countries=list(d.keys()), n_countries=len(d))
            if not d:
                facts["issues"].append("No currency->rate pairs detected.")

        elif e.shape == R.SHAPE_LONG:
            df = E.parse_long(rows)
            facts["n_countries"] = int(df.shape[0])
            facts["products"] = list(map(str, df.columns[:6]))
            if df.empty:
                facts["issues"].append("Long table parsed to zero rows.")

    except Exception as ex:                       # parsing blew up
        facts["parsed_ok"] = False
        facts["issues"].append(f"Parser error: {type(ex).__name__}: {ex}")

    return facts


def _is_empty(rows: List[List[Any]]) -> bool:
    for r in rows or []:
        for c in r:
            if c is not None and str(c).strip() != "" and str(c).lower() != "nan":
                return False
    return True


# --------------------------------------------------------------------------- #
#  Main audit
# --------------------------------------------------------------------------- #
def audit(sheets: Dict[str, List[List[Any]]]) -> Tuple[List[dict], dict]:
    resolved, sheet_usage = _resolve(sheets)
    records: List[dict] = []

    for e in R.REGISTRY:
        rec = dict(key=e.key, display=e.display, used_by=", ".join(e.used_by),
                   required=e.required, reference=e.reference,
                   matched_sheet="", shape=e.shape, status="", severity="",
                   n_months=0, month_range="", n_countries=0, countries="",
                   n_products=0, products="", issues=[], note=e.note)

        # fabricated tables never come from the file
        if e.fabricated:
            rec.update(status="FABRICATED", severity="INFO",
                       matched_sheet="(none - generated)")
            rec["issues"].append("Not present in the workbook; pipeline fills "
                                 "dummy data. Provide real data to replace it.")
            records.append(rec)
            continue

        sheet = resolved.get(e.key)
        if sheet is None:
            rec.update(status="MISSING",
                       severity=("ERROR" if e.required else "WARN"))
            rec["issues"].append("No sheet matched any expected name: "
                                 + ", ".join(repr(a) for a in e.aliases))
            records.append(rec)
            continue

        rec["matched_sheet"] = sheet

        # collision: more than one expected table matched this same sheet
        if len(sheet_usage.get(sheet, [])) > 1:
            others = [k for k in sheet_usage[sheet] if k != e.key]
            rec.update(status="COLLISION", severity="ERROR")
            rec["issues"].append(f"Sheet {sheet!r} also matched: "
                                 f"{', '.join(others)} - names are ambiguous.")

        rows = sheets[sheet]
        if _is_empty(rows):
            rec.update(status="EMPTY",
                       severity=("ERROR" if e.required else "WARN"))
            rec["issues"].append(f"Sheet {sheet!r} is present but has no data.")
            records.append(rec)
            continue

        facts = _validate(e, rows)
        rec.update(n_months=facts["n_months"], month_range=facts["month_range"],
                   n_countries=facts["n_countries"],
                   countries=", ".join(map(str, facts["countries"][:8])),
                   n_products=facts["n_products"],
                   products=", ".join(map(str, facts["products"][:6])))
        rec["issues"].extend(facts["issues"])

        if rec["status"] != "COLLISION":
            if not facts["parsed_ok"] or facts["issues"]:
                rec.update(status="SHAPE_ISSUE", severity="WARN")
            else:
                rec.update(status="OK", severity="OK")
        records.append(rec)

    # unknown / extra sheets not tied to any expected table
    known = {resolved[k] for k in resolved if resolved[k]}
    extras = [s for s in sheets if s not in known]
    for s in extras:
        empty = _is_empty(sheets[s])
        records.append(dict(
            key="(extra)", display=f"Unrecognised sheet: {s}",
            used_by="", required=False, reference=False, matched_sheet=s,
            shape="unknown", status="UNKNOWN", severity="WARN",
            n_months=0, month_range="", n_countries=0, countries="",
            n_products=0, products="",
            issues=["Sheet is in the file but maps to no expected table."
                    + (" It is also empty." if empty else "")],
            note=""))

    summary = _summarise(records, extras)
    records.sort(key=lambda r: (SEV_ORDER.get(r["severity"], 9), r["key"]))
    return records, summary


def _summarise(records: List[dict], extras: List[str]) -> dict:
    counts = defaultdict(int)
    for r in records:
        counts[r["status"]] += 1
    required_missing = [r["display"] for r in records
                        if r["required"] and r["status"] in
                        ("MISSING", "EMPTY", "COLLISION")]
    fabricated = [r["display"] for r in records if r["status"] == "FABRICATED"]
    errors = [r for r in records if r["severity"] == "ERROR"]
    warns = [r for r in records if r["severity"] == "WARN"]
    total_expected = len([r for r in records if r["key"] != "(extra)"])
    ok = counts.get("OK", 0)
    return dict(
        total_expected=total_expected, ok=ok,
        errors=len(errors), warnings=len(warns),
        counts=dict(counts),
        required_missing=required_missing, fabricated=fabricated,
        extras=extras,
        verdict=("READY" if not errors else "NEEDS ATTENTION"),
    )


# --------------------------------------------------------------------------- #
#  Console
# --------------------------------------------------------------------------- #
def print_summary(records: List[dict], summary: dict) -> None:
    print("=" * 74)
    print(f"INPUT DIAGNOSTICS  -  verdict: {summary['verdict']}")
    print("=" * 74)
    print(f"Expected tables: {summary['total_expected']}   "
          f"OK: {summary['ok']}   Errors: {summary['errors']}   "
          f"Warnings: {summary['warnings']}")
    if summary["required_missing"]:
        print("\nREQUIRED but missing/empty/ambiguous:")
        for m in summary["required_missing"]:
            print(f"   - {m}")
    if summary["fabricated"]:
        print("\nFabricated (dummy) - supply real data to replace:")
        for m in summary["fabricated"]:
            print(f"   - {m}")
    if summary["extras"]:
        print("\nUnrecognised sheets in the file:")
        for m in summary["extras"]:
            print(f"   - {m}")
    print("\nPer-table:")
    icon = {"OK": "OK  ", "INFO": "INFO", "WARN": "WARN", "ERROR": "ERR "}
    for r in records:
        line = (f"  [{icon.get(r['severity'],'?')}] {r['status']:11} "
                f"{r['display'][:38]:38} <- {r['matched_sheet']}")
        print(line)
        for iss in r["issues"]:
            print(f"          . {iss}")


# --------------------------------------------------------------------------- #
#  Excel report
# --------------------------------------------------------------------------- #
def write_report(records: List[dict], summary: dict, path: str) -> None:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sev_fill = {"OK": "C6EFCE", "INFO": "DDEBF7", "WARN": "FFEB9C",
                "ERROR": "FFC7CE"}
    sev_font = {"OK": "006100", "INFO": "1F4E78", "WARN": "9C6500",
                "ERROR": "9C0006"}

    detail_rows = [{
        "Status": r["status"], "Severity": r["severity"],
        "Table": r["display"], "Matched sheet": r["matched_sheet"],
        "Shape": r["shape"], "# Months": r["n_months"],
        "Month range": r["month_range"], "# Countries": r["n_countries"],
        "Countries (sample)": r["countries"], "# Products": r["n_products"],
        "Products (sample)": r["products"], "Used by": r["used_by"],
        "Issues": " | ".join(r["issues"]), "Note": r["note"],
    } for r in records]
    df = pd.DataFrame(detail_rows)

    summ_rows = [
        ["Verdict", summary["verdict"]],
        ["Expected tables", summary["total_expected"]],
        ["OK", summary["ok"]],
        ["Errors", summary["errors"]],
        ["Warnings", summary["warnings"]],
        ["", ""],
        ["Status breakdown", ""],
    ]
    for k, v in summary["counts"].items():
        summ_rows.append([f"   {k}", v])
    summ_rows += [["", ""], ["Required missing/empty/ambiguous", ""]]
    summ_rows += [["   " + m, ""] for m in summary["required_missing"]] or \
                 [["   (none)", ""]]
    summ_rows += [["", ""], ["Fabricated (dummy - supply real data)", ""]]
    summ_rows += [["   " + m, ""] for m in summary["fabricated"]] or \
                 [["   (none)", ""]]
    summ_rows += [["", ""], ["Unrecognised sheets", ""]]
    summ_rows += [["   " + m, ""] for m in summary["extras"]] or \
                 [["   (none)", ""]]
    sdf = pd.DataFrame(summ_rows, columns=["Metric", "Value"])

    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        sdf.to_excel(xw, sheet_name="Summary", index=False)
        df.to_excel(xw, sheet_name="Table checks", index=False)

    # format
    from openpyxl import load_workbook
    wb = load_workbook(path)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill("solid", fgColor="1F4E78")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)

    # Summary sheet
    ws = wb["Summary"]
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 60
    for c in range(1, 3):
        ws.cell(1, c).fill = hfill
        ws.cell(1, c).font = hfont
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1)
        if a.value and not str(a.value).startswith("   ") and str(a.value) not in ("",):
            a.font = Font(name="Arial", bold=True, size=10)
        else:
            a.font = Font(name="Arial", size=10)
        ws.cell(r, 2).font = Font(name="Arial", size=10)
        v = ws.cell(r, 1).value
        if v == "Verdict":
            good = ws.cell(r, 2).value == "READY"
            ws.cell(r, 2).font = Font(name="Arial", bold=True, size=11,
                                      color=("006100" if good else "9C0006"))

    # Table checks sheet
    ws = wb["Table checks"]
    widths = [12, 9, 34, 26, 20, 9, 22, 11, 34, 10, 30, 30, 50, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).fill = hfill
        ws.cell(1, c).font = hfont
        ws.cell(1, c).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        sev = ws.cell(r, 2).value
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name="Arial", size=9)
            cell.border = border
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(c in (3, 9, 11, 12, 13, 14)))
        if sev in sev_fill:
            for c in (1, 2):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=sev_fill[sev])
                ws.cell(r, c).font = Font(name="Arial", size=9, bold=True,
                                          color=sev_font[sev])
    ws.freeze_panes = "A2"
    wb.save(path)
