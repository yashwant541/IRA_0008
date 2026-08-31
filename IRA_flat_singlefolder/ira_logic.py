"""
ira_logic.py  -  web-app helpers for table views, TRACE downloads and the
full "IRA Calculation Logic" workbook.

Everything here works off objects the analysis already produced and cached
(the parsed `tables`, the per-metric `intermediates`, the final `frames` and the
resolved `per_cat`).  The IRA engine is imported lazily and never modified.
"""
from __future__ import annotations
import io
from typing import Dict, List, Any, Optional

import pandas as pd


# --------------------------------------------------------------------------- #
#  engine handles (lazy, dual import - package or flat)
# --------------------------------------------------------------------------- #
def _cfg():
    try:
        from IRA import ira_config as C
    except Exception:
        import ira_config as C
    return C


PRODUCTS = ["Secured", "Unsecured", "SME Banking", "Wealth Lending",
            "Wealth Lending - Retail Banking", "Wealth Lending - PvB"]


# =========================================================================== #
#  1.  TABLE VIEWS  -  render every parsed table as a simple grid for the UI
# =========================================================================== #
def _months(mt) -> List[str]:
    from_engine = getattr(mt, "months", None) or []
    # prefer the formatted Mon-YY labels when the engine exposes them
    fmt = getattr(mt, "month_labels", None)
    return list(fmt) if fmt else [str(m) for m in from_engine]


def _fmt_month(m) -> str:
    try:
        from IRA import ira_engine as E
    except Exception:
        import ira_engine as E
    try:
        return E.fmt_month(m)
    except Exception:
        return str(m)


def _grid_from_monthtable(mt, name: str) -> Dict[str, Any]:
    months = [ _fmt_month(m) for m in (getattr(mt, "months", []) or []) ]
    pd_data = getattr(mt, "product_data", None) or {}
    cd_data = getattr(mt, "country_data", None) or {}

    def _row_values(series) -> List[Any]:
        # series may be a dict {month: value} or a list aligned to months
        if isinstance(series, dict):
            return [series.get(m) for m in (getattr(mt, "months", []) or [])]
        if isinstance(series, (list, tuple)):
            return list(series)
        return []

    rows = []
    if pd_data:
        cols = ["Country", "Product"] + months
        for (country, product), series in pd_data.items():
            rows.append([country, product] + _row_values(series))
    elif cd_data:
        cols = ["Country"] + months
        for country, series in cd_data.items():
            rows.append([country] + _row_values(series))
    else:
        cols = ["(no rows parsed)"]
    return {"name": name, "columns": cols, "rows": rows}


def _grid_from_df(df: pd.DataFrame, name: str) -> Dict[str, Any]:
    df = df.where(pd.notnull(df), None)
    return {"name": name, "columns": [str(c) for c in df.columns],
            "rows": df.astype(object).values.tolist()}


def table_views(tables: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Return a list of {name, columns, rows} grids for the parsed input tables,
    so the UI can show each table and let the user copy it out."""
    views: List[Dict[str, Any]] = []

    # nice display names for the month tables
    NAMES = {"ENR": "ENR", "30+%": "30+%", "30+$": "30+$", "90+%": "90+%",
             "90+$": "90+$", "RWA": "RWA", "app_rate": "Country prod level app rate",
             "new_approved": "# monthly new approved", "PPI": "PPI",
             "LTV80": "LTV > 80 Excl MIP", "interest_rates": "Interest Rates"}
    for key, label in NAMES.items():
        mt = tables.get(key)
        if mt is not None and hasattr(mt, "months"):
            g = _grid_from_monthtable(mt, label)
            if g["rows"]:
                views.append(g)

    # dict-of-MonthTable tables (policy L2/L3, EA/AWC, PvB EA/AWC)
    for key, label in (("policy_exception", "# policy exception L2 and L3"),
                       ("ME_EA_AWC", "ME EA AWC"), ("PvB_EA_AWC", "PvB EA AWC")):
        d = tables.get(key)
        if isinstance(d, dict):
            for sub, mt in d.items():
                if hasattr(mt, "months"):
                    g = _grid_from_monthtable(mt, f"{label} [{sub}]")
                    if g["rows"]:
                        views.append(g)

    # GCO / CCPL volatile (dicts of series)
    for key, label in (("gco", "GCO %"), ("ccpl_volatile", "CCPL Volatile by Country")):
        d = tables.get(key)
        if isinstance(d, dict) and d:
            rows = []
            for k, v in d.items():
                if isinstance(v, dict):
                    rows.append([k, ", ".join(f"{kk}={vv}" for kk, vv in list(v.items())[:6])])
                else:
                    rows.append([k, v])
            if rows:
                views.append({"name": label, "columns": ["Key", "Value(s)"], "rows": rows})

    # ECL (DataFrame)
    ecl = tables.get("ecl")
    if isinstance(ecl, pd.DataFrame) and not ecl.empty:
        views.append(_grid_from_df(ecl.head(200), "ECL IIP LI"))

    # sovereign -> Country | FCY CRG | LCY CRG | Outlook
    sov = tables.get("sovereign")
    if isinstance(sov, dict) and sov:
        rows = [[c, r.get("fcy_crg"), r.get("lcy_crg"), r.get("outlook")]
                for c, r in sov.items()]
        views.append({"name": "Country Rating & Outlook",
                      "columns": ["Country", "FCY CRG", "LCY CRG", "Outlook"], "rows": rows})

    # dispensations & breaches (dict category -> {country: count})
    for key, label in (("dispensations", "Active or Expired Dispensation"),
                       ("cra_breaches", "CRA Breaches")):
        d = tables.get(key)
        if isinstance(d, dict) and d:
            rows = []
            for cat, cc in d.items():
                if isinstance(cc, dict):
                    for country, val in cc.items():
                        rows.append([cat, country, val])
            if rows:
                views.append({"name": label, "columns": ["Category", "Country", "Value"],
                              "rows": rows})
    return views


# =========================================================================== #
#  2.  RELEVANT VALUES (TRACE)  -  per product, only the values that matter
# =========================================================================== #
def _fmt_val(x) -> str:
    if x is None:
        return ""
    if isinstance(x, float):
        return f"{x:.4f}".rstrip("0").rstrip(".")
    return str(x)


def _relevant_from_comp(int_key: str, rec: dict) -> Dict[str, str]:
    c = rec or {}
    blank = {"Metric": "", "Current Month": "", "Current Value": "",
             "Reference Month": "", "Reference Value": "", "Notes": c.get("reason", "")}
    if int_key in ("enr_yoy", "ppi_yoy"):
        prior_key = next((k for k in c if k.startswith("prior_")), None)
        return {"Metric": "YoY %", "Current Month": c.get("current_month", ""),
                "Current Value": _fmt_val(c.get("current")),
                "Reference Month": c.get("prior_month", ""),
                "Reference Value": _fmt_val(c.get(prior_key) if prior_key else None),
                "Notes": c.get("reason", "") or (f"ppi key={c['ppi_key']}" if "ppi_key" in c else "")}
    if int_key in ("dpd_qoq_cur", "dpd_qoq_prior", "dpd_yoy"):
        ref_lbl = next((k[:-6] for k in c if k.endswith(" Month") and k != "Current Month"), "Reference")
        return {"Metric": "Deterioration (pp)", "Current Month": c.get("Current Month", ""),
                "Current Value": _fmt_val(c.get("Current Month Value")),
                "Reference Month": c.get(f"{ref_lbl} Month", ""),
                "Reference Value": _fmt_val(c.get(f"{ref_lbl} Month Value")),
                "Notes": c.get("reason", "")}
    if int_key == "dpd_pct_total":
        return {"Metric": "Share of total", "Current Month": c.get("month", ""),
                "Current Value": _fmt_val(c.get("product_current")),
                "Reference Month": "all in-scope countries",
                "Reference Value": _fmt_val(c.get("product_total_across_countries")),
                "Notes": c.get("reason", "")}
    if int_key == "policy_exc_rate":
        return {"Metric": "Ratio (12m)", "Current Month": c.get("window", ""),
                "Current Value": _fmt_val(c.get("exceptions_L2+L3_12m")),
                "Reference Month": "new approved (12m)",
                "Reference Value": _fmt_val(c.get("new_approved_12m")),
                "Notes": c.get("reason", "")}
    if int_key in ("ea_prop", "awc_prop"):
        num_lbl = "EA" if "EA" in c else ("AWC" if "AWC" in c else "num")
        den_lbl = "PvB ENR" if "PvB ENR" in c else "ENR"
        return {"Metric": "Ratio to ENR", "Current Month": num_lbl,
                "Current Value": _fmt_val(c.get(num_lbl)), "Reference Month": den_lbl,
                "Reference Value": _fmt_val(c.get(den_lbl)), "Notes": c.get("reason", "")}
    if int_key == "ltv":
        return {**blank, "Metric": "Latest value", "Current Value": _fmt_val(c.get("ltv_raw")),
                "Notes": (c.get("basis", "") + (" " + c["reason"] if c.get("reason") else "")).strip()}
    if int_key == "volatile":
        return {**blank, "Metric": "Latest value",
                "Notes": (f"matched {c.get('matched') or c.get('matched_code','')}".strip() or c.get("reason", ""))}
    if int_key == "interest_inc":
        return {"Metric": "Increase vs 3yr avg", "Current Month": "current",
                "Current Value": _fmt_val(c.get("current")), "Reference Month": "3yr avg",
                "Reference Value": _fmt_val(c.get("avg_3yr")),
                "Notes": c.get("reason", "") or c.get("basis", "")}
    if int_key in ("sovereign_outlook", "sovereign_grade", "dispensations", "breaches"):
        return {**blank, "Metric": "Latest value"}
    return blank


def _final_frame(frames, product) -> pd.DataFrame:
    df = frames.get(f"IRA - {product}")
    if df is None:
        return pd.DataFrame()
    keep = [c for c in ["Period", "Country", "Label", "Value", "Risk Rating", "Risk Number"]
            if c in df.columns]
    return df[keep].copy()


def _period(frames) -> str:
    f = frames.get("IRA - Secured")
    if f is not None and not f.empty and "Period" in f.columns:
        return str(f["Period"].iloc[0])
    return ""


def relevant_values_frame(product, frames, intermediates) -> pd.DataFrame:
    C = _cfg()
    period = _period(frames)
    fin = frames.get(f"IRA - {product}")
    order = []
    if fin is not None and not fin.empty:
        seen = set()
        for c in fin["Country"]:
            if c not in seen and c != "GROUP":
                seen.add(c); order.append(c)
    rows = []
    for country in order:
        fin_c = fin[fin["Country"] == country].set_index("Label")
        for md in C.METRICS[product]():
            int_key = getattr(md["value"], "int_key", "")
            rec = None
            if int_key:
                store = intermediates.get(int_key, {})
                rec = store.get((country, product)) or store.get((country, None))
            rel = _relevant_from_comp(int_key, rec or {})
            frow = fin_c.loc[md["label"]] if md["label"] in fin_c.index else None
            rows.append({
                "Period": period, "Country": country, "Label": md["label"],
                "Metric": rel["Metric"], "Current Month": rel["Current Month"],
                "Current Value": rel["Current Value"], "Reference Month": rel["Reference Month"],
                "Reference Value": rel["Reference Value"],
                "Computed": ("" if frow is None else frow["Value"]),
                "Risk Rating": ("" if frow is None else frow["Risk Rating"]),
                "Risk Number": ("" if frow is None else frow["Risk Number"]),
                "Notes": rel["Notes"]})
    cols = ["Period", "Country", "Label", "Metric", "Current Month", "Current Value",
            "Reference Month", "Reference Value", "Computed", "Risk Rating",
            "Risk Number", "Notes"]
    return pd.DataFrame(rows, columns=cols)


# =========================================================================== #
#  3.  WORKBOOKS  (TRACE only, and the full LOGIC workbook)
# =========================================================================== #
def _styling():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    navy = "1F4E78"
    return {
        "HFONT": Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "TITLE": Font(name="Arial", bold=True, size=14, color=navy),
        "SUB": Font(name="Arial", bold=True, size=11, color=navy),
        "BODY": Font(name="Arial", size=10), "BOLD": Font(name="Arial", size=10, bold=True),
        "WRAP": Alignment(wrap_text=True, vertical="top"),
        "CENTER": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "FILL": PatternFill("solid", fgColor=navy),
        "BORDER": Border(*(Side(style="thin", color="BFBFBF"),) * 4),
        "navy": navy,
    }


RATING_FILL = {"Very Low": "C6EFCE", "Low": "D9EAD3", "Medium": "FFEB9C",
               "High": "FCE4D6", "Very High": "FFC7CE",
               "Not Available": "F2F2F2", "Not Applicable": "F2F2F2"}


def _trace_sheet(wb, product, frames, intermediates, first=False):
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    S = _styling()
    ws = wb.active if first else wb.create_sheet()
    ws.title = ("Trace - " + product)[:31]
    ws["A1"] = f"Trace - {product}"; ws["A1"].font = S["TITLE"]
    ws["A2"] = "Every country: the relevant values used, the computed value, the rating and the risk number."
    ws["A2"].font = S["BODY"]
    heads = ["Country", "Label", "Metric", "Current Month", "Current Value", "Reference Month",
             "Reference Value", "Computed", "Risk Rating", "Risk No.", "Notes"]
    widths = [14, 40, 18, 15, 14, 20, 14, 12, 14, 9, 30]
    for i, h in enumerate(heads):
        c = ws.cell(4, i + 1, h); c.font = S["HFONT"]; c.fill = S["FILL"]
        c.alignment = S["CENTER"]; c.border = S["BORDER"]
        ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]
    rel = relevant_values_frame(product, frames, intermediates)
    r = 5
    for _, row in rel.iterrows():
        rating = str(row["Risk Rating"])
        vals = [row["Country"], row["Label"], row["Metric"], row["Current Month"],
                row["Current Value"], row["Reference Month"], row["Reference Value"],
                row["Computed"], rating, ("" if str(row["Risk Number"]) == "nan" else row["Risk Number"]),
                row["Notes"]]
        for i, v in enumerate(vals):
            cell = ws.cell(r, i + 1, v); cell.font = S["BODY"]
            cell.alignment = S["WRAP"]; cell.border = S["BORDER"]
        ws.cell(r, 9).fill = PatternFill("solid", fgColor=RATING_FILL.get(rating, "FFFFFF"))
        r += 1
    ws.freeze_panes = "A5"


def _group_trace_sheet(wb, frames, tables):
    """One 'GROUP calc trace' sheet: the full arithmetic behind every GROUP row."""
    if tables is None:
        return
    try:
        from . import ira_group as G
    except ImportError:
        import ira_group as G
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    S = _styling()
    ws = wb.create_sheet("GROUP calc trace")
    ws["A1"] = "GROUP calc trace - the full arithmetic behind every Country='GROUP' row"
    ws["A1"].font = S["TITLE"]
    ws["A2"] = ("Table-operation labels sum over ALL countries; ENR-weighted labels list each "
                "config country's (risk number x ENR weight). Weight = country ENR / total ENR over all countries.")
    ws["A2"].font = S["BODY"]
    heads = ["Product / Label", "Detail", "Value", "Rating", "Risk No."]
    widths = [44, 46, 16, 12, 9]
    for i, h in enumerate(heads):
        c = ws.cell(4, i + 1, h); c.font = S["HFONT"]; c.fill = S["FILL"]
        c.alignment = S["CENTER"]; c.border = S["BORDER"]
        ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]
    try:
        trace = G.group_trace(frames, tables)
    except Exception as e:
        ws.cell(5, 1, "GROUP trace unavailable: " + str(e)).font = S["BODY"]
        return
    r = 5
    for product in PRODUCTS:
        entries = trace.get(product)
        if not entries:
            continue
        cell = ws.cell(r, 1, "GROUP - " + product); cell.font = S["SUB"]
        for cc in range(1, 6):
            ws.cell(r, cc).fill = S["FILL"]; ws.cell(r, cc).font = S["HFONT"]; ws.cell(r, cc).border = S["BORDER"]
        ws.cell(r, 1, "GROUP - " + product)
        r += 1
        for e in entries:
            ws.cell(r, 1, e["label"]).font = S["BOLD"]
            ws.cell(r, 2, e["kind"]).font = S["BODY"]
            ws.cell(r, 3, e["value"]).font = S["BOLD"]
            rt = str(e["rating"] or "")
            rc = ws.cell(r, 4, rt); rc.font = S["BOLD"]
            rc.fill = PatternFill("solid", fgColor=RATING_FILL.get(rt, "FFFFFF"))
            ws.cell(r, 5, (e["number"] if e["number"] not in (None, "") else "")).font = S["BOLD"]
            for cc in range(1, 6):
                ws.cell(r, cc).border = S["BORDER"]; ws.cell(r, cc).alignment = S["WRAP"]
            r += 1
            for what, val in e["detail"]:
                ws.cell(r, 2, what).font = S["BODY"]
                ws.cell(r, 3, val).font = S["BODY"]
                for cc in range(1, 6):
                    ws.cell(r, cc).border = S["BORDER"]; ws.cell(r, cc).alignment = S["WRAP"]
                r += 1
        r += 1
    ws.freeze_panes = "A5"


def build_trace_workbook(frames, intermediates, tables=None) -> bytes:
    """A workbook with one Trace sheet per product (relevant values + ratings),
    plus a 'GROUP calc trace' sheet when tables are provided."""
    from openpyxl import Workbook
    wb = Workbook()
    for i, p in enumerate(PRODUCTS):
        _trace_sheet(wb, p, frames, intermediates, first=(i == 0))
    _group_trace_sheet(wb, frames, tables)
    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()


def build_logic_workbook(frames, intermediates, tables=None) -> bytes:
    """The full 'IRA Calculation Logic' workbook: the reference sheets (pipeline,
    input files, label logic, rating ladders, final score) plus the live
    worked-example and per-product Trace sheets - all from this run's data."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    S = _styling()
    wb = Workbook()

    def hdr(ws, row, headers, widths=None, start=1):
        for i, h in enumerate(headers):
            c = ws.cell(row, start + i, h); c.font = S["HFONT"]; c.fill = S["FILL"]
            c.alignment = S["CENTER"]; c.border = S["BORDER"]
        if widths:
            for i, w in enumerate(widths):
                ws.column_dimensions[get_column_letter(start + i)].width = w

    def row(ws, r, values, fonts=None, fills=None, start=1):
        for i, v in enumerate(values):
            c = ws.cell(r, start + i, v); c.font = (fonts[i] if fonts else S["BODY"])
            c.alignment = S["WRAP"]; c.border = S["BORDER"]
            if fills and fills[i]:
                c.fill = PatternFill("solid", fgColor=fills[i])

    # ---- Read me ----
    ws = wb.active; ws.title = "Read me"
    ws["A1"] = "IRA - Inherent Credit Risk Assessment - calculation logic"; ws["A1"].font = S["TITLE"]
    ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 116
    read = [
        ("The pipeline (zero to 100)", "sub"),
        ("Three inputs feed the assessment: the MI file, the Other-tables file and the countries config.", "b"),
        ("Step 1 EXTRACT - each raw sheet is tidied (Country / Product / months) with a Period = latest month.", ""),
        ("Step 2 RELEVANT VALUES - keep only the current month & value and the reference month & value (or", ""),
        ("        numerator & denominator) per label - the transparent, checkable layer (see the Trace sheets).", ""),
        ("Step 3 VALUE - compute the metric (YoY %, deterioration in pp, a ratio, or a point-in-time reading).", ""),
        ("Step 4 RISK RATING - pass the value through that label's ladder -> Very Low .. Very High.", ""),
        ("Step 5 RISK NUMBER - Very Low=1, Low=2, Medium=3, High=4, Very High=5 (blank if N/A).", ""),
        ("Step 6 CALCULATED - group the labels, take each group's worst number, weight and sum -> the rating.", ""),
        ("", ""),
        ("Step 7 GROUP - each product also gets Country='GROUP' rows: table operations", ""),
        ("        (ratio-of-totals over ALL countries) for the deterioration/policy/LTV/", ""),
        ("        volatile/EA/AWC labels; ENR country-% weight for the rest (see '4b').", ""),
        ("", ""),
        ("Sheets: '1. Input files', '2. Label logic', '3. Rating ladders', '4. Final score',", "b"),
        ("'4b. GROUP roll-up', a live worked example, and one 'Trace - <Product>' per product.", "b"),
    ]
    r = 3
    for t, k in read:
        c = ws.cell(r, 2, t); c.font = {"sub": S["SUB"], "b": S["BOLD"]}.get(k, S["BODY"]); r += 1

    # ---- 1. Input files ----
    ws = wb.create_sheet("1. Input files")
    ws["A1"] = "1. Input files -> tables -> labels"; ws["A1"].font = S["TITLE"]
    hdr(ws, 3, ["Input file", "Raw table / sheet", "Feeds labels"], [20, 40, 60])
    inp = [
        ("MI file", "ENR", "1a (all); ENR is the EA/AWC denominator"),
        ("MI file", "90+% / 30+%", "1bi, 1bii, 1c deterioration (Secured 90+, others 30+)"),
        ("MI file", "90+$ / 30+$", "1d share of DPD$ total; GROUP 1bi/1bii/1c ratio-of-totals (all countries)"),
        ("MI file", "# policy exception L2 and L3 ; # monthly new approved", "policy rate (Sec/Unsec 1e, SME 1g, Wealth 1g)"),
        ("MI file", "ME EA AWC / PvB EA AWC", "EA & AWC proportions (SME 1e/1f, Wealth 1d/1e)"),
        ("MI file", "LTV > 80 Excl MIP", "Secured 1g"),
        ("MI file", "CCPL Volatile by Country", "Unsecured 1g"),
        ("MI file", "Interest Rates", "Secured 2a (engine fallback if absent)"),
        ("Other tables", "Property Price Index", "Secured 2b"),
        ("Other tables", "Country Rating (FCY CRG) / Country Outlook", "grading (2d/2b) & outlook (2c/2a)"),
        ("Other tables", "Active or Expired Dispensation", "dispensations (Sec 1f, Unsec 1f, SME 1h, Wealth 1h)"),
        ("Other tables", "Credit Risk Appetite Breaches", "breaches in last 12m (Sec 1h, Unsec 1h, SME 1i)"),
        ("Countries config", "Category / Country / Include", "which countries run; the in-scope set 1d sums over"),
    ]
    rr = 4
    for v in inp:
        row(ws, rr, v, fonts=[S["BOLD"], S["BODY"], S["BODY"]]); rr += 1

    # ---- 2. Label logic + 3. Rating ladders + 4. Final score (reference text) ----
    _reference_sheets(wb, S, hdr, row)

    # ---- 5. Worked example (live, first country present) ----
    ws = wb.create_sheet("5. Worked example")
    ws["A1"] = "5. Worked example - end to end"; ws["A1"].font = S["TITLE"]
    example_country = None
    fsec = frames.get("IRA - Secured")
    if fsec is not None and not fsec.empty:
        example_country = "UAE" if (fsec["Country"] == "UAE").any() else fsec["Country"].iloc[0]
    ws["A2"] = f"Country shown: {example_country}"; ws["A2"].font = S["BODY"]
    r = 4
    for p in PRODUCTS:
        rel = relevant_values_frame(p, frames, intermediates)
        sub = rel[rel["Country"] == example_country]
        if sub.empty:
            continue
        ws.cell(r, 1, f"{p}  -  {example_country}").font = S["SUB"]; r += 1
        heads = ["Label", "Metric", "Current Month", "Current Value", "Reference Month",
                 "Reference Value", "Computed", "Rating", "No.", "Notes"]
        hdr(ws, r, heads, [42, 18, 15, 14, 20, 14, 12, 12, 6, 28]); r += 1
        for _, rw in sub.iterrows():
            rating = str(rw["Risk Rating"])
            row(ws, r, [rw["Label"], rw["Metric"], rw["Current Month"], rw["Current Value"],
                        rw["Reference Month"], rw["Reference Value"], rw["Computed"], rating,
                        ("" if str(rw["Risk Number"]) == "nan" else rw["Risk Number"]), rw["Notes"]])
            ws.cell(r, 8).fill = PatternFill("solid", fgColor=RATING_FILL.get(rating, "FFFFFF"))
            r += 1
        fin = _final_frame(frames, p)
        calc = fin[(fin["Country"] == example_country) & (fin["Label"].str.startswith("Calculated"))]
        if not calc.empty:
            cr = calc.iloc[0]
            row(ws, r, ["Calculated Inherent Credit Risk Assessment:", "", "", "", "", "",
                        cr["Value"], str(cr["Risk Rating"]), "", ""], fonts=[S["BOLD"]] * 10)
            r += 1
        r += 1

    # ---- Trace sheets (live) ----
    for p in PRODUCTS:
        _trace_sheet(wb, p, frames, intermediates, first=False)
    _group_trace_sheet(wb, frames, tables)

    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()


def _reference_sheets(wb, S, hdr, row):
    """Static reference sheets: label logic, rating ladders, final score."""
    from openpyxl.styles import PatternFill
    # 2. Label logic (concise master)
    ws = wb.create_sheet("2. Label logic")
    ws["A1"] = "2. Label logic - how each label's Value is calculated"; ws["A1"].font = S["TITLE"]
    hdr(ws, 3, ["Product", "ID", "Label", "Value - what is calculated", "Rating ladder"],
        [16, 6, 40, 60, 26])
    YOY = "YoY % = (current - value 12m back)/|12m back|"
    DETc = "DPD% at current - DPD% 3m back (pp)"; DETp = "DPD% 1m back - DPD% 4m back (pp)"
    DETy = "DPD% at current - DPD% 12m back (pp)"; SHARE = "country DPD$ / sum DPD$ over in-scope countries"
    POL = "(L2+L3 over 12m) / (new approved over 12m)"; EA = "latest EA $ / latest ENR"
    AWC = "latest AWC $ / latest ENR"; LTV = "latest LTV>80 / 100"
    VOL = "country's own CCPL code value (2/3-letter); Global is the GROUP value only"
    INT = "(current rate - 36m avg)/100"; PPI = "YoY % of Property Price Index"
    OUT = "Outlook text (Positive/Stable/Negative)"; GRD = "FCY CRG grade (leading number)"
    DISP = "count of active dispensations"; BRCH = "breaches over last 12 months (L12M)"
    BL = "blank for this product -> Not Applicable"
    data = {
        "Secured": [("1a", "Asset Growth YoY %", YOY, "ENR-YoY (Secured)"),
                    ("1bi", "QoQ Deterioration 90+ current", DETc, "Deterioration pair (AND)"),
                    ("1bii", "QoQ Deterioration 90+ prior", DETp, "Deterioration pair (AND)"),
                    ("1c", "YoY Deterioration 90+", DETy, "Deterioration"),
                    ("1d", "Country 90+ share", SHARE, "1d 90+ share (Secured)"),
                    ("1e", "Policy exceptions rate", POL, "Policy (Secured)"),
                    ("1f", "Active dispensations", DISP, "Dispensations"),
                    ("1g", "LTV>80 concentration", LTV, "LTV>80"),
                    ("1h", "Breaches last 12m", BRCH, "Breaches"),
                    ("2a", "Interest rate increase", INT, "Interest"),
                    ("2b", "YoY change in PPI", PPI, "PPI (higher better)"),
                    ("2c", "Country outlook", OUT, "Outlook"),
                    ("2d", "Country grading", GRD, "Grading")],
        "Unsecured": [("1a", "Asset Growth YoY %", YOY, "ENR-YoY (others)"),
                      ("1bi", "QoQ Deterioration 30+ current", DETc, "Deterioration pair (AND)"),
                      ("1bii", "QoQ Deterioration 30+ prior", DETp, "Deterioration pair (AND)"),
                      ("1c", "YoY Deterioration 30+", DETy, "Deterioration"),
                      ("1d", "Country 30+ share", SHARE, "1d 30+ share (Unsecured)"),
                      ("1e", "Policy exceptions rate", POL, "Policy (Unsecured)"),
                      ("1f", "Active dispensations", DISP, "Dispensations"),
                      ("1g", "Volatile segment concentration", VOL, "Volatile"),
                      ("1h", "Breaches last 12m", BRCH, "Breaches"),
                      ("2a", "Country outlook", OUT, "Outlook"),
                      ("2b", "Country grading", GRD, "Grading")],
        "SME Banking": [("1a", "Asset Growth YoY %", YOY + " on SME Banking + ME ENR combined", "ENR-YoY (others)"),
                        ("1bi", "QoQ Deterioration 30+ current", DETc, "Deterioration pair (AND)"),
                        ("1bii", "QoQ Deterioration 30+ prior", DETp, "Deterioration pair (AND)"),
                        ("1c", "YoY Deterioration 30+", DETy, "Deterioration"),
                        ("1d", "Country 30+ share", SHARE, "1d 30+ share (SME)"),
                        ("1e", "EA proportion", "latest EA $ / latest ENR (Product = ME)", "EA (SME)"),
                        ("1f", "AWC proportion", "latest AWC $ / latest ENR (Product = ME)", "AWC (SME)"),
                        ("1g", "Policy exceptions rate", POL, "Policy (SME/Wealth)"),
                        ("1h", "Active dispensations", DISP, "Dispensations"),
                        ("1i", "Breaches last 12m", BRCH, "Breaches"),
                        ("2a", "Country outlook", OUT, "Outlook"),
                        ("2b", "Country grading", GRD, "Grading")],
        "Wealth Lending / Retail / PvB": [
            ("1a", "Asset Growth YoY %", YOY + " (total = Wealth+PvB; Retail = Wealth; PvB = PvB)", "ENR-YoY (others)"),
            ("1bi/ii,1c", "Deterioration 30+", "per-country blank for PvB; GROUP uses 30+$ Wealth Banking", "Deterioration"),
            ("1d", "EA proportion", "latest EA $ / latest ENR (Product = PvB); blank for Retail", "EA/AWC (Wealth)"),
            ("1e", "AWC proportion", "latest AWC $ / latest ENR (Product = PvB); blank for Retail", "EA/AWC (Wealth)"),
            ("1f", "Shortfall", "(securities Total + real-estate Total for country)/1000 / ENR Wealth Banking; Wealth+PvB only, blank for Retail", "Shortfall (Wealth)"),
            ("1g", "Policy exceptions rate", POL, "Policy (SME/Wealth)"),
            ("1h", "Active dispensations", DISP, "Dispensations"),
            ("1i", "Breaches last 12m (Wealth Lending only)", BRCH, "Breaches"),
            ("2a", "Country outlook", OUT, "Outlook"),
            ("2b", "Country grading", GRD, "Grading")],
    }
    r = 4
    for prod, items in data.items():
        for it in items:
            row(ws, r, [prod, it[0], it[1], it[2], it[3]],
                fonts=[S["BOLD"], S["BOLD"], S["BODY"], S["BODY"], S["BODY"]]); r += 1

    # 3. Rating ladders
    ws = wb.create_sheet("3. Rating ladders")
    ws["A1"] = "3. Rating ladders - value takes the first bound it exceeds (fractions: 0.05 = 5%)"
    ws["A1"].font = S["TITLE"]
    hdr(ws, 3, ["Ladder", "Applies to", "Bands (high -> low)"], [26, 26, 80])
    ladders = [
        ("ENR-YoY (Secured)", "Secured 1a", ">.10 VeryHigh | >.05 High | >.03 Medium | >.01 Low | else VeryLow"),
        ("ENR-YoY (others)", "others 1a", ">.15 VeryHigh | >.05 High | >.03 Medium | >.01 Low | else VeryLow"),
        ("Deterioration pair - Secured", "Secured 1bi/1bii", "both cur AND prior > : >.0005 VH | >.0003 H | >.0002 M | >.0001 L | else VL"),
        ("Deterioration single - Secured", "Secured 1c", ">.001 VH | >.0005 H | >.0003 M | >.0001 L | else VL"),
        ("Deterioration pair - Unsec/SME", "Unsec/SME 1bi/1bii", "both > : >.0025 VH | >.0006 H | >.0004 M | >.0001 L | else VL"),
        ("Deterioration single - Unsec/SME", "Unsec/SME 1c", ">.0025 VH | >.002 H | >.0015 M | >.001 L | else VL"),
        ("Deterioration pair - Wealth", "Wealth 1bi/1bii", "both > : >.0025 VH | >0 H | else VL"),
        ("Deterioration single - Wealth", "Wealth 1c", ">.0025 VH | >0 H | else VL"),
        ("1d 90+ share (Secured)", "Secured 1d", ">.25 VH | >.15 H | >.05 M | >.025 L | else VL"),
        ("1d 30+ share (Unsecured)", "Unsecured 1d", ">.10 VH | >.075 H | >.05 M | >.025 L | else VL"),
        ("1d 30+ share (SME)", "SME 1d", ">.20 VH | >.15 H | >.10 M | >.05 L | else VL"),
        ("Policy (Secured)", "Secured 1e", ">.15 VH | >.10 H | >.075 M | >.05 L | else VL"),
        ("Policy (Unsecured)", "Unsecured 1e", ">.05 VH | >.01 H | >.005 M | >.0025 L | else VL"),
        ("Policy (SME)", "SME 1g", ">.075 VH | >.05 H | >.03 M | >.01 L | else VL"),
        ("Policy (Wealth)", "Wealth 1g", ">.035 VH | >.025 H | >.015 M | >.005 L | else VL"),
        ("EA (SME)", "SME 1e", ">.10 VH | >.075 H | >.05 M | >.025 L | else VL"),
        ("AWC (SME)", "SME 1f", ">.125 VH | >.10 H | >.075 M | >.05 L | else VL"),
        ("EA (Wealth)", "Wealth 1d", ">.01 VH | >.0075 H | >.005 M | >.0025 L | else VL"),
        ("AWC (Wealth)", "Wealth 1e", ">.035 VH | >.025 H | >.015 M | >.005 L | else VL"),
        ("Shortfall (Wealth)", "Wealth 1f", ">.25 VH | >.10 H | >.05 M | >.03 L | else VL"),
        ("LTV>80", "Secured 1g", ">.10 VH | >.05 H | >.025 M | >.01 L | else VL"),
        ("Volatile", "Unsecured 1g", ">.125 VH | >.10 H | >.075 M | >.05 L | else VL"),
        ("Interest", "Secured 2a", ">.03 VH | >.02 H | >.01 M | >0 L | else VL"),
        ("PPI (higher better)", "Secured 2b", ">.05 VeryLow | >0 Low | >-.05 Medium | >-.15 High | else VeryHigh"),
        ("Dispensations (count)", "1f / 1h", ">3 VH | =3 H | =2 M | =1 L | 0 VL ; missing -> Not Available"),
        ("Breaches (L12M count)", "1h / 1i", ">1 VH | =1 High | 0 VeryLow ; missing -> Not Available"),
        ("Outlook", "2c / 2a", "Positive VeryLow | Stable Low | Negative VeryHigh"),
        ("Grading (exact grades)", "2d / 2b", "VH:11A/B/C,12A/B/C,13 | H:8B,9A,9B,10A | M:5B,6A,6B,7A,7B,8A | VL:1A,1B,2A | else Low"),
    ]
    r = 4
    for v in ladders:
        row(ws, r, v, fonts=[S["BOLD"], S["BODY"], S["BODY"]]); r += 1

    # 4. Final score
    ws = wb.create_sheet("4. Final score")
    ws["A1"] = "4. Calculated Inherent Credit Risk Assessment"; ws["A1"].font = S["TITLE"]
    txt = [
        ("Each group contributes its worst (max) risk number; score = SUM over groups of (1/6 x max).", "b"),
        ("score >= 4.5 Very High | >= 3.5 High | >= 2.5 Medium | >= 1.5 Low | else Very Low", "b"),
        ("(comparison uses the un-rounded score; a value that displays 1.5 but is 1.4999.. rates Very Low).", ""),
        ("", ""),
        ("Groups per product (label positions):", "sub"),
    ]
    r = 3
    for t, k in txt:
        ws.cell(r, 1, t).font = {"sub": S["SUB"], "b": S["BOLD"]}.get(k, S["BODY"]); r += 1
    groups = [("Secured", "[1] [2,3,4,5] [6,7,8] [9] [10] [11,12,13]"),
              ("Unsecured", "[1] [2,3,4,5] [6,7] [8] [9] [10,11]"),
              ("SME Banking", "[1] [2,3,4,5,6,7] [8,9] [10] [11,12]"),
              ("Wealth (all three)", "[1] [2,3,4,5,6,7] [8,9] [10,11]")]
    hdr(ws, r, ["Product", "Groups"], [24, 60]); r += 1
    for v in groups:
        row(ws, r, v, fonts=[S["BOLD"], S["BODY"]]); r += 1

    # 4b. GROUP (portfolio) roll-up
    ws = wb.create_sheet("4b. GROUP roll-up")
    ws["A1"] = "4b. GROUP roll-up - the Country = 'GROUP' rows added to each product"
    ws["A1"].font = S["TITLE"]
    intro = [
        ("Each product output carries a block of GROUP rows: one per label plus a GROUP", "b"),
        ("'Calculated Inherent Credit Risk Assessment:' row. Each label is one of two kinds.", "b"),
        ("", ""),
        ("KIND 1 - TABLE OPERATIONS: a ratio of totals over ALL countries in the raw table", "sub"),
        ("(not the countries-config scope), rated with the SAME ladder the per-country label", ""),
        ("uses. Which labels are table operations, per product:", ""),
        ("   Secured   : 1bi 1bii 1c , 1e policy , 1g LTV", ""),
        ("   Unsecured : 1bi 1bii 1c , 1e policy , 1g volatile", ""),
        ("   SME       : 1bi 1bii 1c , 1e EA , 1f AWC , 1g policy", ""),
        ("   Wealth L. : 1bi 1bii 1c , 1d EA , 1e AWC , 1f shortfall , 1g policy", ""),
        ("   Retail/PvB: 1bi 1bii 1c (+ 1f shortfall for PvB); rest are weighted", ""),
        ("   Wealth Lending 1i (breaches) is ENR-weighted, not a table operation", ""),
    ]
    r = 3
    for t, k in intro:
        ws.cell(r, 1, t).font = {"sub": S["SUB"], "b": S["BOLD"]}.get(k, S["BODY"]); r += 1
    hdr(ws, r, ["Label", "GROUP value = ", "Notes"], [30, 60, 46]); r += 1
    ratio = [
        ("1bi QoQ deterioration (current)",
         "DPD%(current) - DPD%(current - 1 quarter)",
         "DPD%(m) = SUM(DPD$)/SUM(category ENR), all countries; Secured 90+$, others 30+$"),
        ("1bii QoQ deterioration (prior)",
         "DPD%(prior month) - DPD%(prior month - 1 quarter)",
         "shown as its own row; excluded from the GROUP inherent (1bi drives the pair)"),
        ("1c YoY deterioration",
         "DPD%(current) - DPD%(current - 1 year)",
         "all three Wealth products use the 30+$ 'Wealth Banking' line"),
        ("policy exceptions rate",
         "SUM(L2+L3 over 12m, all) / SUM(new approved over 12m, all)",
         "Secured/Unsec 1e, SME/Wealth 1g"),
        ("LTV>80 (Secured 1g)",
         "the LTV>80 table's own 'Total' row (raw fraction)",
         "point-in-time, latest month"),
        ("volatile (Unsecured 1g)",
         "the CCPL table's portfolio 'Global' value",
         "GROUP only; per-country reads each country's own 2/3-letter code"),
        ("EA proportion (SME 1e, Wealth 1d)",
         "SUM(EA$, all countries, latest) / SUM(ENR, latest); ENR: SME=ME, Wealth=PvB",
         "numerator: SME = ME EA/AWC table, Wealth = PvB EA/AWC table; denom ENR: SME=ME, Wealth=PvB"),
        ("AWC proportion (SME 1f, Wealth 1e)",
         "SUM(AWC$, all countries, latest) / SUM(ENR, latest)",
         "table op for SME & Wealth Lending; Retail/PvB EA-AWC are ENR-weighted"),
        ("shortfall (Wealth 1f)",
         "(securities Total-Amount + real-estate Total-Amount)/1000 / SUM(ENR Wealth Banking)",
         "Wealth Lending & PvB; Retail blank; 1i breaches is ENR-weighted, not here"),
    ]
    for v in ratio:
        row(ws, r, v, fonts=[S["BOLD"], S["BODY"], S["BODY"]]); r += 1
    r += 1
    tail = [
        ("KIND 2 - ENR COUNTRY-% WEIGHT (every other label).", "sub"),
        ("weight(country) = that country's category ENR (latest month) / TOTAL category ENR", ""),
        ("over ALL countries (latest month) - NOT renormalised, so the config countries'", ""),
        ("weights need not sum to 1. GROUP number = SUM over the product's config countries", ""),
        ("of (per-country risk number x weight), rounded to the nearest 1..5; then its rating.", ""),
        ("The three Wealth products share the Wealth Lending ENR (Wealth Banking + PvB).", ""),
        ("", ""),
        ("Each GROUP row is tagged in the output ('GROUP table operation (all countries)' or", ""),
        ("'GROUP ENR-weighted (country %)') so you can see which rule produced it.", ""),
        ("", ""),
        ("GROUP inherent: same theme groups as sheet 4; take the worst (max) GROUP risk", "sub"),
        ("number in each theme, x the theme weight, summed - with 1bii excluded.", ""),
    ]
    for t, k in tail:
        ws.cell(r, 1, t).font = {"sub": S["SUB"], "b": S["BOLD"]}.get(k, S["BODY"]); r += 1
