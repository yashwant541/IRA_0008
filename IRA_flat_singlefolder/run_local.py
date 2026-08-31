"""
run_local.py
============
Runs the IRA engine against a local Excel and writes, in this order:

  1) IRA_Normalized_Inputs.xlsx  - cleaned block sources in LONG form
                                    (Country | Product | Mar-25 ... Mar-26),
                                    output product names, one sheet per source.
  2) IRA_ByCategory.xlsx         - four per-category tables per core source
                                    (e.g. ENR-Secured), Mar-26 headers.
  3) IRA_Intermediate.xlsx       - Mapping (per category) + every calculated
                                    metric table (with reasons).
  4) IRA_Output.xlsx             - the four final IRA tables (with NA reasons).

Intermediates are produced BEFORE the finals.

Usage:
    python run_local.py INPUT.xlsx OUTPUT.xlsx
Countries per category are read from countries_config.csv (auto-seeded).
"""

import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- make imports work no matter where you run this from -------------------- #
# Put both this script's folder AND its IRA/ subfolder on the path, so either
# layout works:  (a) run_local.py next to the IRA/ package, or
#                (b) run_local.py next to the loose ira_*.py modules.
_HERE = os.path.dirname(os.path.abspath(__file__))
for _p in (_HERE, os.path.join(_HERE, "IRA")):
    if os.path.isdir(_p) and _p not in sys.path:
        sys.path.insert(0, _p)

try:
    from IRA import (ira_loaders as L, ira_build as B, ira_normalize as N,
                     ira_engine as E, ira_countries as CC, ira_io as IO,
                     ira_preview as PV)
except ImportError:
    try:
        import ira_loaders as L, ira_build as B, ira_normalize as N
        import ira_engine as E, ira_countries as CC, ira_io as IO
        import ira_preview as PV
    except ImportError as ex:
        sys.exit(
            "ERROR: could not import the IRA modules (%s).\n"
            "Expected either:\n"
            "  - an 'IRA' folder (with ira_engine.py, ira_build.py, ...) next to "
            "this script, or\n"
            "  - the ira_*.py files directly next to this script.\n"
            "This script is at: %s\n"
            "Folders/files seen here: %s" % (
                ex, _HERE, ", ".join(sorted(os.listdir(_HERE)))))


# Editable input list controlling which countries each category is run for.
COUNTRIES_CONFIG = "countries_config.csv"


RATING_FILL = {"Very Low": "C6EFCE", "Low": "D9EAD3", "Medium": "FFEB9C",
               "High": "FCE4D6", "Very High": "FFC7CE", "Not Available": "F2F2F2"}
RATING_FONT = {"Very Low": "006100", "Low": "375623", "Medium": "9C6500",
               "High": "833C00", "Very High": "9C0006", "Not Available": "808080"}
HFILL = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
WARN = PatternFill("solid", fgColor="FFEB9C")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read_local(path):
    """Read forcing cached formula values (input files are full of formulas)."""
    sheets, warns = IO.read_workbook_checked(path)
    for w in warns:
        print("   [reader]", w)
    return sheets


def _write_sheets(path, frames: dict, first=None):
    """Write {title: df} to an xlsx, sanitising sheet names (no / \\ : * ? [ ])."""
    used = set()
    order = list(frames.items())
    if first and first in frames:
        order = [(first, frames[first])] + [(k, v) for k, v in order if k != first]
    safe = []
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for title, df in order:
            name = E.sanitize_sheet_name(title, used)
            safe.append(name)
            (df if not df.empty else pd.DataFrame({"(empty)": []})).to_excel(
                xw, sheet_name=name, index=False)
    _format(path, safe)


def _format(path, sheet_names):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for c in range(1, ws.max_column + 1):
            ws.cell(1, c).fill = HFILL
            ws.cell(1, c).font = HFONT
            ws.cell(1, c).alignment = Alignment(vertical="center", wrap_text=True)
            w = 22
            h = str(headers[c-1] or "")
            if h in ("Label", "Calculation", "Source table(s)", "Reason",
                     "What to do in Value Column", "Intermediate table"):
                w = 40
            elif h in ("Country", "Product", "Category", "Metric"):
                w = 16
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"
        # rating colouring
        if "Risk Rating" in headers:
            rc = headers.index("Risk Rating") + 1
            for r in range(2, ws.max_row + 1):
                rv = ws.cell(r, rc).value
                if rv in RATING_FILL:
                    ws.cell(r, rc).fill = PatternFill("solid", fgColor=RATING_FILL[rv])
                    ws.cell(r, rc).font = Font(name="Arial", size=10, bold=True,
                                               color=RATING_FONT[rv])
        # reason highlighting
        for col_name in ("Reason", "What to do in Value Column"):
            if col_name in headers:
                rc = headers.index(col_name) + 1
                for r in range(2, ws.max_row + 1):
                    if ws.cell(r, rc).value:
                        ws.cell(r, rc).fill = WARN
    wb.save(path)


def main():
    inp = sys.argv[1] if len(sys.argv) > 1 else "Dummy.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "IRA_Output.xlsx"
    base = out[:-5] if out.lower().endswith(".xlsx") else out

    # Fail fast: the country list is user-defined; the script never creates it.
    if not os.path.exists(COUNTRIES_CONFIG):
        sys.exit(
            f"ERROR: '{COUNTRIES_CONFIG}' not found.\n"
            "This file is user-defined and must be created by you (the script "
            "will not create it).\n"
            "It needs columns:  Category,Country,Include  (Include = Yes/No),\n"
            "with one row per (category, country) you want included.\n"
            "Categories: Secured, Unsecured, SME Banking, Wealth Lending.\n"
            "The IRA runs ONLY for the countries marked Include=Yes per category.")

    sheets = read_local(inp)

    # Merge any additional reference-table workbooks (dispensations, breaches,
    # sovereign, PPI, interest rates) passed as extra args or found alongside.
    import glob
    extra_files = list(sys.argv[3:]) if len(sys.argv) > 3 else []
    _dir = os.path.dirname(os.path.abspath(inp)) or "."
    for pat in ("Other_Tables.xlsx", "*Other*.xlsx",
                "*reference*.xlsx", "*Reference*.xlsx"):
        for f in glob.glob(os.path.join(_dir, pat)):
            if os.path.abspath(f) != os.path.abspath(inp) and f not in extra_files:
                extra_files.append(f)
    for f in extra_files:
        try:
            for nm, rows in read_local(f).items():
                sheets[f"{os.path.basename(f)}::{nm}"] = rows
            print(f"Merged reference tables from {os.path.basename(f)}")
        except Exception as ex:
            print(f"Could not read {f}: {ex}")

    # ---- 0) DETECTION PREVIEW - prove every sheet was understood -------- #
    summary, cleaned = PV.build_preview(sheets)
    prev = {"Detection Summary": summary, **cleaned}
    _write_sheets("IRA_Detection_Preview.xlsx", prev, first="Detection Summary")
    ok = int((summary["Status"] == "OK").sum())
    print(f"[0] IRA_Detection_Preview.xlsx  ({ok}/{len(summary)} sheets detected "
          f"as monthly tables)")
    for _, r in summary.iterrows():
        flag = "OK " if r["Status"] == "OK" else "!! "
        print(f"    {flag}{r['Sheet']:<28} {r['Status']:<16} "
              f"{r['# Date columns']} date cols, {r['# Data rows']} rows"
              + (f"  [{r['Note']}]" if r["Note"] else ""))

    tables = L.load_tables(sheets)

    # ---- country selection: STRICTLY from the user-provided config -------- #
    countries_per_category = CC.load(COUNTRIES_CONFIG)
    if not countries_per_category:
        sys.exit(f"ERROR: '{COUNTRIES_CONFIG}' has no rows marked Include=Yes. "
                 "Add the countries you want per category and re-run.")
    per_cat = B.resolve_countries(tables, None, countries_per_category)
    print("Countries per category (from countries_config.csv):")
    for k, v in per_cat.items():
        print(f"   {k:14} {len(v)} -> {v}")

    # ---- 1) normalized inputs (LONG) ---------------------------------- #
    long_frames = N.normalized_long(tables)
    _write_sheets("IRA_Normalized_Inputs.xlsx", long_frames)
    print(f"\n[1] IRA_Normalized_Inputs.xlsx  ({len(long_frames)} long tables)")

    # ---- 2) per-category tables --------------------------------------- #
    cat_frames = N.normalized_by_category(tables)
    _write_sheets("IRA_ByCategory.xlsx", cat_frames)
    print(f"[2] IRA_ByCategory.xlsx         ({len(cat_frames)} category tables)")

    # ---- 3) calculated intermediates + mapping ------------------------ #
    inter = B.build_intermediate_frames(tables, per_cat)
    mapping = B.build_mapping()
    inter_all = {"Mapping": mapping, **inter}
    _write_sheets("IRA_Intermediate.xlsx", inter_all, first="Mapping")
    print(f"[3] IRA_Intermediate.xlsx       (Mapping + {len(inter)} calc tables)")

    # ---- 4) final outputs --------------------------------------------- #
    frames = B.build_all(tables, countries_per_category=countries_per_category)
    na = 0
    for name, df in frames.items():
        n = int((df["Risk Rating"] == "Not Available").sum())
        na += n
        print(f"      {name:22} {len(df):>3} rows, {n} Not Available")
    _write_sheets(out, frames)
    print(f"[4] {out}   (total Not-Available cells: {na})")

    # NA reasons summary
    print("\nNot-Available reasons:")
    seen = set()
    for name, df in frames.items():
        for _, row in df[df["Risk Rating"] == "Not Available"].iterrows():
            msg = str(row["What to do in Value Column"])
            if msg and msg not in seen:
                seen.add(msg); print(f"   - {msg}")


if __name__ == "__main__":
    main()
